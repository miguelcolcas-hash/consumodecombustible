import streamlit as st
import requests
import zipfile
import io
import pandas as pd
from datetime import datetime, timedelta, date
import urllib.parse
import plotly.express as px
import openpyxl

# --- 1. CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Supervisión YUPANA e IEOD - Osinergmin", layout="wide")
st.title("🛢️ Dashboard Integral - Consumo de Combustible (SEIN)")
st.markdown("### Fiscalización Dinámica: Lo Programado (YUPANA) vs. Lo Ejecutado (IEOD)")

# --- 2. PARÁMETROS OPERATIVOS COMUNES ---
MES_TXT = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SETIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
MESES_IEOD = {i+1: MES_TXT[i].capitalize() for i in range(12)}

# --- 3. ETL YUPANA (PROGRAMADO) ---
archivos_clave_yupana = {
    "TERMICA"      : "Termica - Despacho (MW)",
    "COMBUSTIBLE"  : "Termica - Consumo de Combustible"
}

def clasificar_tecnologia_yupana(nombre_central):
    nombre = str(nombre_central).upper()
    diesel_kws = ["D2", "R6", "RESIDUAL", "DIESEL", "ILO21", "ILO 21", "ILO1", "ILO 1", "MOLLENDO", "RECKA", "INDEPENDENCIA", "SAMANCO", "TARAPOTO", "IQUITOS", "YURIMAGUAS", "PUERTO MALDONADO", "BELLAVISTA", "PEDRO RUIZ", "ETEN", "PIURA D", "CALANA", "ELOR", "SHCUMMINS", "SNTV", "NEPI", "PUERTO BRAVO", "NODO"]
    if any(kw in nombre for kw in diesel_kws): return "Residual+Diésel D2"
    return "Otra Tecnología"

def cargar_df_desde_zip(zf, stem):
    for info in zf.infolist():
        nombre_base = info.filename.split('/')[-1]
        if stem in nombre_base and not nombre_base.startswith("~"):
            with zf.open(info) as f:
                if nombre_base.upper().endswith('.CSV'): 
                    try: return pd.read_csv(f, sep=None, engine='python')
                    except:
                        f.seek(0)
                        return pd.read_csv(f, sep=',')
    return None

def extraer_todas_centrales(df):
    series = {}
    if df is None or df.empty: return series
    invalid_cols = ["HORA", "TIEMPO", "FECHA", "ETAPA", "GENERADOR"]
    if df.shape[1] > 1:
        cols = [c for c in df.columns if not any(inv in str(c).upper() for inv in invalid_cols) and not str(c).startswith("Unnamed")]
        for c in cols: series[c] = pd.to_numeric(df[c], errors='coerce').fillna(0).tolist()
    else:
        enc = [h.strip() for h in str(df.columns[0]).split(",")]
        start_idx = 0
        if len(enc) < 2:
            enc = [h.strip() for h in str(df.iloc[0,0]).split(",")]
            start_idx = 1
        nombres_validos, idx_validos = [], []
        for i, nombre in enumerate(enc[1:], start=1):
            if not any(inv in nombre.upper() for inv in invalid_cols):
                nombres_validos.append(nombre)
                idx_validos.append(i)
                series[nombre] = []
        for fila in df.iloc[start_idx:, 0].astype(str):
            partes = [p.strip() for p in fila.split(",")]
            for nombre, i in zip(nombres_validos, idx_validos):
                series[nombre].append(float(partes[i]) if i < len(partes) and partes[i] else 0.0)
    return series

def extraer_motivo_dinamico(y, m, M, d, ddmm, l, headers):
    urls = [
        f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{l}%2FReprog_{ddmm}{l}.xlsx",
        f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog_{ddmm}{l}%2FReprog_{ddmm}{l}.xlsx"
    ]
    for u in urls:
        try:
            r = requests.get(u, headers=headers, timeout=10)
            if r.status_code == 200 and len(r.content) > 1000:
                wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
                ws = wb.worksheets[0]
                for row in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=3).value
                    if cell_value and "MOTIVO" in str(cell_value).upper():
                        motivo_val = ws.cell(row=row+1, column=4).value
                        if motivo_val: return str(motivo_val).strip()
                return "No se encontró motivo."
        except:
            pass
    return "No se pudo extraer origen."

def rellenar_hasta_48(lst):
    if not lst: return [0.0]*48
    faltan = 48 - len(lst)
    return ([0.0]*faltan + lst) if faltan > 0 else lst[:48]

def suma_elementos_variable(*listas):
    if not listas: return []
    length = max(len(l) for l in listas if l)
    if length == 0: return []
    out = [0.0]*length
    for lst in listas:
        if lst:
            for i in range(min(length, len(lst))):
                if pd.notna(lst[i]): out[i] += lst[i]
    return out

@st.cache_data(show_spinner=False, ttl=300)
def extraer_datos_yupana_memoria(f):
    y, m, d = f.strftime("%Y"), f.strftime("%m"), f.strftime("%d")
    M = MES_TXT[f.month-1]
    fecha_str = f"{y}{m}{d}"
    ddmm = f"{d}{m}"
    
    headers = {'User-Agent': 'Mozilla/5.0'}
    datos_dia = {"Dataframes": {}, "Log": []}
    
    # PDO
    url_pdo = f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FPrograma%20Diario%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FYUPANA_{fecha_str}.zip"
    try:
        r = requests.get(url_pdo, headers=headers, timeout=15)
        if r.status_code == 200 and r.content[:4] == b'PK\x03\x04':
            with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
                datos_dia["Dataframes"]["PDO"] = {}
                for key, stem in archivos_clave_yupana.items():
                    datos_dia["Dataframes"]["PDO"][key] = extraer_todas_centrales(cargar_df_desde_zip(zf, stem))
            datos_dia["Log"].append("✅ PDO")
        else: datos_dia["Log"].append("❌ PDO")
    except Exception: datos_dia["Log"].append("❌ PDO")

    # RDO Dinámico
    letra_actual = 'A'
    while True:
        nombre_rdo = f"RDO_{letra_actual}"
        urls_rdo = [
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{letra_actual}%2FYUPANA_{ddmm}{letra_actual}.zip",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog_{ddmm}{letra_actual}%2FYUPANA_{ddmm}{letra_actual}.zip",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}%20{letra_actual}%2FYUPANA_{ddmm}{letra_actual}.zip"
        ]
        
        exito_rdo = False
        for enlace in urls_rdo:
            if exito_rdo: break
            try:
                r = requests.get(enlace, headers=headers, timeout=10)
                if r.status_code == 200 and r.content[:4] == b'PK\x03\x04':
                    with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
                        datos_dia["Dataframes"][nombre_rdo] = {}
                        for key, stem in archivos_clave_yupana.items():
                            datos_dia["Dataframes"][nombre_rdo][key] = extraer_todas_centrales(cargar_df_desde_zip(zf, stem))
                    datos_dia["Dataframes"][f"MOTIVO_{nombre_rdo}"] = extraer_motivo_dinamico(y, m, M, d, ddmm, letra_actual, headers)
                    datos_dia["Log"].append(f"✅ {nombre_rdo}")
                    exito_rdo = True
            except Exception: continue 
        
        if exito_rdo: letra_actual = chr(ord(letra_actual) + 1)
        else:
            datos_dia["Log"].append(f"🛑 Fin en {nombre_rdo}")
            break
            
    return datos_dia

def crear_grafica_area_apilada(df_plot, marcadores=None):
    df_plot = df_plot.copy().fillna(0)
    num_cols = [c for c in df_plot.columns if c != 'Hora']
    df_plot[num_cols] = df_plot[num_cols].apply(pd.to_numeric, errors='coerce').fillna(0).round(2)
    
    # Cálculo del total del sistema por cada hora
    df_plot['TOTAL_SISTEMA'] = df_plot[num_cols].sum(axis=1).round(2)
    totales_por_unidad = df_plot.drop(columns=['Hora', 'TOTAL_SISTEMA']).sum()
    orden_columnas = totales_por_unidad.sort_values(ascending=False).index.tolist()
    
    # Melt para la gráfica apilada
    cols_mantener = ['Hora'] + orden_columnas
    df_melt = df_plot[cols_mantener].melt(id_vars=['Hora'], var_name='Unidad Generadora', value_name='Consumo')
    
    fig = px.area(df_melt, x="Hora", y="Consumo", color="Unidad Generadora", labels={"Consumo": "Consumo Físico"})
    fig.update_xaxes(tickformat="%d/%m %H:%M", tickangle=45)
    
    # LÍNEA INVISIBLE PARA MOSTRAR EL TOTAL EN EL HOVER
    fig.add_scatter(
        x=df_plot['Hora'], 
        y=df_plot['TOTAL_SISTEMA'], 
        mode='lines', 
        line=dict(width=0, color='rgba(0,0,0,0)'), 
        name='<b>⚡ TOTAL CONSUMO</b>', 
        showlegend=False
    )
    
    for trace in fig.data:
        trace.hoverinfo = ['skip' if pd.isna(v) or float(v) <= 0.01 else 'all' for v in trace.y]
        if 'TOTAL CONSUMO' in trace.name: 
            trace.hovertemplate = '<b>%{y:,.2f} Galones</b><br>%{x|%d/%m %H:%M}'
        else: 
            trace.hovertemplate = "%{y:,.2f} Galones"
    
    if marcadores:
        for ts, texto in marcadores:
            fig.add_vline(x=ts, line_width=1.5, line_dash="dash", line_color="rgba(255,255,255,0.7)")
            align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
            fig.add_annotation(x=ts, y=1.02, yref="paper", text=f"<b>{texto} {ts.strftime('%H:%M')}</b>", showarrow=False, font=dict(size=10, color="white"), bgcolor="#e74c3c", bordercolor="white", borderwidth=1, borderpad=3, textangle=-90, yanchor="bottom", xanchor=align)
            
    fig.update_layout(hovermode="x unified", height=600, margin=dict(t=120, b=50, l=60, r=50), yaxis_title="Consumo de Combustible (Galones)")
    return fig

# --- 4. ETL IEOD (EJECUTADO / POST-OPERACIÓN) ---
def generar_urls_coes(fecha):
    año = fecha.strftime("%Y")
    mes_num = fecha.strftime("%m")
    dia = fecha.strftime("%d")
    mes_titulo = MESES_IEOD[fecha.month]
    fecha_str = fecha.strftime("%d%m")
    
    path_nuevo = f"Post Operación/Reportes/IEOD/{año}/{mes_num}_{mes_titulo}/{dia}/AnexoA_{fecha_str}.xlsx"
    path_legacy = f"Post Operación/Reportes/IEOD/{año}/{mes_num}_{mes_titulo}/{dia}/Anexo1_Resumen_{fecha_str}.xlsx"
    return [
        (f"https://www.coes.org.pe/portal/browser/download?url={urllib.parse.quote(path_nuevo)}", "AnexoA"),
        (f"https://www.coes.org.pe/portal/browser/download?url={urllib.parse.quote(path_legacy)}", "Anexo1")
    ]

@st.cache_data(show_spinner=False)
def extraer_datos_ieod(fecha):
    urls = generar_urls_coes(fecha)
    headers = {'User-Agent': 'Mozilla/5.0'}
    df_raw = None
    
    for url, tipo_anexo in urls:
        try:
            res = requests.get(url, headers=headers, timeout=20)
            if res.status_code == 200:
                archivo_excel = io.BytesIO(res.content)
                xls = pd.ExcelFile(archivo_excel, engine='openpyxl')
                hojas_limpias = {h.strip().upper(): h for h in xls.sheet_names}
                
                if "CONSUMO_COMB" in hojas_limpias:
                    df_raw = pd.read_excel(xls, sheet_name=hojas_limpias["CONSUMO_COMB"], header=6, usecols="B:G")
                    break
        except Exception: continue
            
    if df_raw is None or df_raw.empty: return pd.DataFrame(), f"[{fecha.strftime('%d/%m/%Y')}] No se halló IEOD."

    try: df_raw.columns = ['EMPRESA', 'CENTRAL', 'MEDIDOR', 'TIPO_COMBUSTIBLE', 'UNIDAD_MEDIDA', 'CONSUMO']
    except ValueError as e: return pd.DataFrame(), f"[{fecha.strftime('%d/%m/%Y')}] Error estructura COES."

    df_raw = df_raw.dropna(subset=['EMPRESA', 'CENTRAL'])
    df_raw['CONSUMO'] = df_raw['CONSUMO'].astype(str).str.replace(',', '', regex=False)
    df_raw['CONSUMO'] = pd.to_numeric(df_raw['CONSUMO'], errors='coerce').fillna(0)
    
    if df_raw.empty: return pd.DataFrame(), f"[{fecha.strftime('%d/%m/%Y')}] Tabla vacía."

    for col in ['EMPRESA', 'CENTRAL', 'MEDIDOR', 'TIPO_COMBUSTIBLE', 'UNIDAD_MEDIDA']:
        df_raw[col] = df_raw[col].astype(str).str.strip().str.upper()

    mask_gas = df_raw['TIPO_COMBUSTIBLE'].str.contains('GAS', na=False)
    mask_m3 = df_raw['UNIDAD_MEDIDA'].str.contains('M3', na=False)
    df_raw.loc[mask_gas & mask_m3, 'CONSUMO'] = df_raw.loc[mask_gas & mask_m3, 'CONSUMO'] / 1000000.0
    df_raw.loc[mask_gas & mask_m3, 'UNIDAD_MEDIDA'] = 'Mm3'
    
    mask_diesel = df_raw['TIPO_COMBUSTIBLE'].str.contains('DIESEL|RESIDUAL', na=False)
    df_raw.loc[mask_diesel, 'UNIDAD_MEDIDA'] = 'Galones'
    
    df_raw['FECHA_OPERATIVA'] = pd.to_datetime(fecha)
    
    return df_raw, None

# --- 5. INTERFAZ Y EJECUCIÓN ---
st.sidebar.header("Parámetros de Fiscalización")
rango_fechas = st.sidebar.date_input("Intervalo de Fechas", value=(date.today() - timedelta(days=2), date.today() - timedelta(days=1)))

st.sidebar.markdown("### Acciones de Extracción")
btn_extraer = st.sidebar.button("⚡ Extraer Datos (YUPANA e IEOD)", type="primary", use_container_width=True)

if btn_extraer:
    if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
        ini, fin = rango_fechas
        status, prog_bar = st.empty(), st.progress(0)
        
        # --- EXTRACCIÓN YUPANA ---
        log_exp = st.expander("Ver bitácora de descargas YUPANA", expanded=False)
        datos_completos_yupana = {}
        dias = (fin - ini).days + 1
        
        for k in range(dias):
            f_actual = ini + timedelta(days=k)
            status.markdown(f"**⏳ [1/2] Vectorizando YUPANA (Programado):** {f_actual.strftime('%d/%m/%Y')}")
            datos_dia = extraer_datos_yupana_memoria(f_actual)
            datos_completos_yupana[f_actual] = datos_dia
            with log_exp: st.markdown(f"**{f_actual.strftime('%d/%m/%Y')}** ➔ " + " | ".join(datos_dia["Log"]))
            prog_bar.progress(((k + 1) / dias) * 0.5)
            
        st.session_state['datos_yupana'] = datos_completos_yupana
        
        # --- EXTRACCIÓN IEOD ---
        fechas_ieod = pd.date_range(ini, fin)
        total_dias_ieod = len(fechas_ieod)
        lista_dfs_ieod, alertas_ieod = [], []
        
        for i, f in enumerate(fechas_ieod):
            status.markdown(f"**⏳ [2/2] Procesando IEOD (Ejecutado):** {f.strftime('%d/%m/%Y')}")
            df_dia, error = extraer_datos_ieod(f)
            if not df_dia.empty: lista_dfs_ieod.append(df_dia)
            if error: alertas_ieod.append(error)
            prog_bar.progress(0.5 + (((i + 1) / total_dias_ieod) * 0.5))
                
        if lista_dfs_ieod: st.session_state['df_ieod'] = pd.concat(lista_dfs_ieod, ignore_index=True)
        else: st.session_state['df_ieod'] = pd.DataFrame()
        
        st.session_state['alertas_ieod'] = alertas_ieod
        
        status.success("✅ Motores YUPANA e IEOD Compilados con Éxito.")
        prog_bar.empty()

st.markdown("---")

# --- 6. VISUALIZACIÓN MULTI-PESTAÑA ---
t_yupana, t_ieod, t_motivos = st.tabs([
    "📅 YUPANA (Programado)", 
    "✅ IEOD (Ejecutado)", 
    "📋 Motivos RDO"
])

# ====== TAB 1: YUPANA (PROGRAMADO) ======
with t_yupana:
    st.info("**Contexto Osinergmin:** Los datos presentados aquí reflejan el **Consumo Programado** de combustible (Reserva Fría / Diésel), extraído de los Programas Diarios (PDO) y Reprogramas (RDO). Para recalcular los totales de forma interactiva, utiliza el filtro 'Filtrar Nodos YUPANA' en lugar de la leyenda.")
    st.markdown("### 📅 Consumo de Combustible PROGRAMADO")
    
    if 'datos_yupana' in st.session_state:
        data = st.session_state['datos_yupana']
        fechas_ordenadas = sorted(data.keys())
        active_prog_dict, ts_dict, dics_cache_dict, marcadores_globales = {}, {}, {}, []
        
        for f in fechas_ordenadas:
            df_dia_sel = data[f]["Dataframes"]
            progs = ["PDO"] if "PDO" in df_dia_sel else []
            progs.extend(sorted([p for p in df_dia_sel.keys() if p.startswith("RDO_")]))
            if not progs: continue
            
            dics_cache = {p: df_dia_sel[p] for p in progs}
            active_prog = [progs[0]] * 48
            if len(progs) > 1:
                for p in progs[1:]:
                    tot = [0.0]*48
                    for v in dics_cache[p].get("TERMICA", {}).values(): tot = suma_elementos_variable(tot, rellenar_hasta_48(v))
                    for i, val in enumerate(tot):
                        if val > 50:
                            for j in range(i, 48): active_prog[j] = p
                            break
                            
            ts_dia = [datetime.combine(f, datetime.min.time()) + timedelta(minutes=30*(i+1)) for i in range(48)]
            p_actual = active_prog[0]
            marcadores_globales.append((ts_dia[0], p_actual))
            for i in range(1, 48):
                if active_prog[i] != p_actual:
                    p_actual = active_prog[i]
                    marcadores_globales.append((ts_dia[i], p_actual))
                    
            active_prog_dict[f], ts_dict[f], dics_cache_dict[f] = active_prog, ts_dia, dics_cache

        dfs_comb = []
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog, dics_cache = active_prog_dict[f], dics_cache_dict[f]
            dia_data_comb = {}
            for i in range(48):
                p = active_prog[i]
                if "COMBUSTIBLE" in dics_cache[p] and dics_cache[p]["COMBUSTIBLE"]:
                    for central, v_list in dics_cache[p]["COMBUSTIBLE"].items():
                        c_clean = str(central).strip()
                        if clasificar_tecnologia_yupana(c_clean) == "Residual+Diésel D2":
                            if c_clean not in dia_data_comb: dia_data_comb[c_clean] = [0.0] * 48
                            dia_data_comb[c_clean][i] += rellenar_hasta_48(v_list)[i]
                            
            df_dia_comb = pd.DataFrame(dia_data_comb)
            df_dia_comb.insert(0, 'Hora', ts_dict[f])
            dfs_comb.append(df_dia_comb)
            
        if dfs_comb:
            df_total_comb = pd.concat(dfs_comb, ignore_index=True).fillna(0)
            num_cols = [c for c in df_total_comb.columns if c != 'Hora']
            active_cols = [c for c in num_cols if df_total_comb[c].sum() > 0]
            
            if not active_cols:
                st.success("✅ No hubo consumo PROGRAMADO de Diésel/Reserva Fría en YUPANA.")
            else:
                todas_centrales_comb = sorted(active_cols)
                filtro_comb = st.multiselect("⚙️ Filtrar Nodos YUPANA:", options=todas_centrales_comb, default=[])
                lista_filtro_comb = filtro_comb if filtro_comb else todas_centrales_comb
                
                df_daily = df_total_comb.copy()
                df_daily['Fecha_Operativa'] = (df_daily['Hora'] - pd.Timedelta(minutes=1)).dt.date
                df_daily_grouped = df_daily.groupby('Fecha_Operativa')[lista_filtro_comb].sum()
                
                daily_totals = df_daily_grouped.sum(axis=1)
                max_centrals = df_daily_grouped.idxmax(axis=1)
                max_vals = df_daily_grouped.max(axis=1)
                
                st.markdown("#### 📊 Consumo Total Diario (Galones)")
                metric_cols = st.columns(len(daily_totals))
                for idx, (fecha_val, total_val) in enumerate(daily_totals.items()):
                    top_unidad = max_centrals[fecha_val]
                    top_val = max_vals[fecha_val]
                    with metric_cols[idx]:
                        st.metric(
                            label=f"📅 {fecha_val.strftime('%d/%m/%Y')} (Prog.)", 
                            value=f"{total_val:,.2f} Gal.",
                            delta=f"🔥 Max: {top_unidad} ({top_val:,.2f} Gal.)",
                            delta_color="off"
                        )
                st.markdown("---")
                
                df_plot_comb = df_total_comb[['Hora'] + lista_filtro_comb]
                st.plotly_chart(crear_grafica_area_apilada(df_plot_comb, marcadores=marcadores_globales), use_container_width=True)
    else: 
        st.warning("👈 Por favor, configura las fechas y haz clic en **'⚡ Extraer Datos'** en el panel lateral.")

# ====== TAB 2: IEOD (EJECUTADO) ======
with t_ieod:
    st.success("**Contexto Osinergmin:** Los datos aquí mostrados representan el **Consumo Ejecutado (Real)**, extraído del IEOD Anexo A/1. Para excluir variables y ver el impacto matemático en tiempo real, utiliza los filtros inferiores.")
    st.markdown("### ✅ Consumo de Combustible EJECUTADO")
    
    if 'df_ieod' in st.session_state:
        df_datos = st.session_state['df_ieod']
        alertas = st.session_state['alertas_ieod']
        
        if df_datos.empty:
            st.error("🚨 Sin datos IEOD en el rango de fechas. (Puede que el reporte real aún no haya sido publicado por el COES).")
            if alertas:
                for a in alertas: st.write(a)
        else:
            if alertas:
                with st.expander("⚠️ Alertas IEOD"):
                    for a in alertas: st.warning(a)
                    
            st.markdown("#### 🔍 Filtrar Variables Ejecutadas")
            c1, c2, c3, c4 = st.columns(4)
            lista_empresas = sorted(df_datos['EMPRESA'].unique())
            with c1: filtro_emp = st.multiselect("🏢 Empresa:", options=lista_empresas)
            
            df_t1 = df_datos[df_datos['EMPRESA'].isin(filtro_emp)] if filtro_emp else df_datos
            with c2: filtro_cen = st.multiselect("⚡ Central:", options=sorted(df_t1['CENTRAL'].unique()))
            
            df_t2 = df_t1[df_t1['CENTRAL'].isin(filtro_cen)] if filtro_cen else df_t1
            with c3: filtro_med = st.multiselect("📟 Medidor:", options=sorted(df_t2['MEDIDOR'].unique()))
            
            df_t3 = df_t2[df_t2['MEDIDOR'].isin(filtro_med)] if filtro_med else df_t2
            with c4: filtro_comb = st.multiselect("🛢️ Combustible:", options=sorted(df_t3['TIPO_COMBUSTIBLE'].unique()))

            df_filtrado = df_datos.copy()
            if filtro_emp: df_filtrado = df_filtrado[df_filtrado['EMPRESA'].isin(filtro_emp)]
            if filtro_cen: df_filtrado = df_filtrado[df_filtrado['CENTRAL'].isin(filtro_cen)]
            if filtro_med: df_filtrado = df_filtrado[df_filtrado['MEDIDOR'].isin(filtro_med)]
            if filtro_comb: df_filtrado = df_filtrado[df_filtrado['TIPO_COMBUSTIBLE'].isin(filtro_comb)]

            if not df_filtrado.empty:
                for comb in sorted(df_filtrado['TIPO_COMBUSTIBLE'].unique()):
                    df_p = df_filtrado[df_filtrado['TIPO_COMBUSTIBLE'] == comb]
                    if df_p.empty: continue
                    unidad = df_p['UNIDAD_MEDIDA'].iloc[0]
                    
                    st.markdown(f"#### 📊 Consumo Total Diario - {comb} ({unidad})")
                    
                    df_grp_kpi = df_p.groupby(['FECHA_OPERATIVA', 'CENTRAL'])['CONSUMO'].sum().reset_index()
                    if not df_grp_kpi.empty and df_grp_kpi['CONSUMO'].sum() > 0:
                        df_grp_kpi_p = df_grp_kpi[df_grp_kpi['CONSUMO'] > 0]
                        
                        totales_dia = df_grp_kpi.groupby('FECHA_OPERATIVA')['CONSUMO'].sum()
                        idx_max = df_grp_kpi_p.groupby('FECHA_OPERATIVA')['CONSUMO'].idxmax()
                        max_centrals = df_grp_kpi_p.loc[idx_max].set_index('FECHA_OPERATIVA')['CENTRAL']
                        max_vals = df_grp_kpi_p.loc[idx_max].set_index('FECHA_OPERATIVA')['CONSUMO']
                        
                        metric_cols = st.columns(len(totales_dia))
                        for idx, (fecha_val, total_val) in enumerate(totales_dia.items()):
                            top_unidad = max_centrals.get(fecha_val, "N/A")
                            top_val = max_vals.get(fecha_val, 0.0)
                            
                            delta_txt = f"🔥 Max: {top_unidad} ({top_val:,.2f} {unidad})" if top_unidad != "N/A" else "N/A"
                            
                            with metric_cols[idx]:
                                st.metric(
                                    label=f"📅 {fecha_val.strftime('%d/%m/%Y')} (Ejec.)",
                                    value=f"{total_val:,.2f} {unidad}",
                                    delta=delta_txt,
                                    delta_color="off"
                                )
                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    opt1, opt2 = st.columns(2)
                    with opt1: mostrar_tot = st.toggle(f"Mostrar Etiqueta de Total Real ({comb})", value=True)
                    with opt2: mostrar_max = st.toggle(f"Mostrar Etiqueta de Central Máxima ({comb})", value=True)
                    
                    df_grp = df_p.groupby(['FECHA_OPERATIVA', 'CENTRAL'])['CONSUMO'].sum().reset_index()
                    df_grp_p = df_grp[df_grp['CONSUMO'] > 0].copy()
                    df_tot = df_grp.groupby('FECHA_OPERATIVA', as_index=False).agg(TOTAL=('CONSUMO', 'sum'))
                    df_max = df_grp_p.loc[df_grp_p.groupby('FECHA_OPERATIVA')['CONSUMO'].idxmax(), ['FECHA_OPERATIVA', 'CENTRAL']].rename(columns={'CENTRAL': 'MAX_CEN'})
                    df_anot = pd.merge(df_tot, df_max, on='FECHA_OPERATIVA', how='left')
                    
                    fig = px.bar(
                        df_grp_p, x="FECHA_OPERATIVA", y="CONSUMO", color="CENTRAL",
                        title=f"Volumen Real Ejecutado de {comb} ({unidad})", text_auto='.4f'
                    )
                    
                    max_y = df_tot['TOTAL'].max()
                    fig.update_layout(height=550, yaxis=dict(range=[0, max_y * 1.25 if max_y>0 else 1]), xaxis_title="Día Operativo", yaxis_title=f"Consumo Real ({unidad})")
                    
                    for _, row in df_anot.iterrows():
                        if row['TOTAL'] > 0 and pd.notna(row['MAX_CEN']):
                            txts = []
                            if mostrar_tot: txts.append(f"<b>Total: {row['TOTAL']:,.4f}</b>")
                            if mostrar_max: txts.append(f"⚡ Max: {row['MAX_CEN']}")
                            if txts: fig.add_annotation(x=row['FECHA_OPERATIVA'], y=row['TOTAL'], text="<br>".join(txts), showarrow=False, yshift=25)
                    
                    fig.update_xaxes(dtick="86400000", tickformat="%d/%m/%Y")
                    st.plotly_chart(fig, use_container_width=True)
                    st.markdown("---")
                    
                st.markdown("#### 🗄️ Trazabilidad de Registros Crudos IEOD")
                df_m = df_filtrado[['FECHA_OPERATIVA', 'EMPRESA', 'CENTRAL', 'MEDIDOR', 'TIPO_COMBUSTIBLE', 'UNIDAD_MEDIDA', 'CONSUMO']].copy()
                df_m['FECHA_OPERATIVA'] = df_m['FECHA_OPERATIVA'].dt.strftime('%d/%m/%Y')
                st.dataframe(df_m, use_container_width=True, hide_index=True, column_config={"CONSUMO": st.column_config.NumberColumn(format="%.6f")})
            else: st.warning("No hay datos reales para la combinación de filtros.")
    else: 
        st.warning("👈 Por favor, configura las fechas y haz clic en **'⚡ Extraer Datos'** en el panel lateral.")

# ====== TAB 3: MOTIVOS ======
with t_motivos:
    st.info("**Contexto Osinergmin:** Documentación de las justificaciones técnicas entregadas por el COES ante los cambios en el consumo programado.")
    st.markdown("### 📋 Motivos de Reprogramación Operativa")
    
    if 'datos_yupana' in st.session_state:
        tabla_motivos = []
        for f in fechas_ordenadas:
            if f not in dics_cache_dict: continue
            for p in sorted(set(active_prog_dict[f])):
                if "RDO" in p:
                    tabla_motivos.append({"Fecha": f.strftime("%d/%m/%Y"), "Reprograma": p, "Motivo Declarado": st.session_state['datos_yupana'][f]["Dataframes"].get(f"MOTIVO_{p}", "Motivo no disponible.")})
        if tabla_motivos: st.dataframe(pd.DataFrame(tabla_motivos), use_container_width=True)
        else: st.success("Sin reprogramas justificados en el periodo extraído.")
    else: 
        st.warning("👈 Por favor, configura las fechas y haz clic en **'⚡ Extraer Datos'** en el panel lateral.")