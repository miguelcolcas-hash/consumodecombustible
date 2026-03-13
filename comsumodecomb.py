import streamlit as st
import requests
import zipfile
import io
import pandas as pd
from datetime import datetime, timedelta, date
import urllib.parse
import plotly.express as px
import openpyxl

# --- 1. CONFIGURACIÓN DE PÁGINA ---
st.set_page_config(page_title="Supervisión YUPANA e IEOD - Osinergmin", layout="wide")
st.title("🛢️ Dashboard Integral - Consumo de Combustible (SEIN)")
st.markdown("### Fiscalización Dinámica: Lo Programado (YUPANA) vs. Lo Ejecutado (IEOD)")

# --- 2. PARÁMETROS OPERATIVOS COMUNES ---
MES_TXT = ["ENERO","FEBRERO","MARZO","ABRIL","MAYO","JUNIO","JULIO","AGOSTO","SETIEMBRE","OCTUBRE","NOVIEMBRE","DICIEMBRE"]
MESES_IEOD = {i+1: MES_TXT[i].capitalize() for i in range(12)}
DIAS_ESP = {0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes', 5: 'Sábado', 6: 'Domingo'}

# --- FUNCIONES AUXILIARES Y FORMATO K/M ---
def formato_k_m(num):
    if pd.isna(num): return "0.00"
    abs_num = abs(num)
    if abs_num >= 1e6:
        return f"{num/1e6:.2f}M"
    elif abs_num >= 1e3:
        return f"{num/1e3:.2f}k"
    else:
        return f"{num:.2f}"

def convertir_volumen(serie_val, serie_unidad, unidad_destino):
    val_m3 = serie_val.copy().astype(float)
    serie_unidad = serie_unidad.astype(str).str.upper()
    
    mask_gal = serie_unidad.str.contains('GAL', na=False)
    mask_bbl = serie_unidad.str.contains('BBL|BARRIL', na=False)
    
    val_m3[mask_gal] = val_m3[mask_gal] / 264.172
    val_m3[mask_bbl] = val_m3[mask_bbl] / 6.28981
    
    if unidad_destino == 'm3': return val_m3
    elif unidad_destino == 'Galones': return val_m3 * 264.172
    elif unidad_destino == 'bbl': return val_m3 * 6.28981
    return val_m3

def agregar_totales_diarios(fig, df, col_fecha, col_valor, prefijo="Total", unidad="", col_tipo=None, barmode="relative"):
    if col_tipo and col_tipo in df.columns:
        df_tot = df.groupby(col_fecha, as_index=False).agg(TOTAL=(col_valor, 'sum'), TIPO=(col_tipo, 'first'))
    else:
        df_tot = df.groupby(col_fecha, as_index=False).agg(TOTAL=(col_valor, 'sum'))
        df_tot['TIPO'] = ""

    if barmode == "group":
        df_max = df.groupby(col_fecha, as_index=False).agg(MAX_Y=(col_valor, 'max'))
        df_tot = pd.merge(df_tot, df_max, on=col_fecha)
    else:
        df_tot['MAX_Y'] = df_tot['TOTAL']

    max_y_global = df_tot['MAX_Y'].max()
    fig.update_layout(yaxis=dict(range=[0, max_y_global * 1.35 if max_y_global > 0 else 1]))
    
    str_unidad = f" {unidad}" if unidad else ""
    for _, row in df_tot.iterrows():
        if row['TOTAL'] > 0:
            val_tipo = str(row.get('TIPO', ''))
            if 'Proyectado' in val_tipo or 'Estimación' in val_tipo:
                tipo_str = " (Proy.)"
            elif 'Ejecutado' in val_tipo:
                tipo_str = " (Ejec.)"
            else:
                tipo_str = ""
                
            lbl_val = f"{row['TOTAL']:,.1f}"
            fig.add_annotation(
                x=row[col_fecha], y=row['MAX_Y'], 
                text=f"<b>{prefijo}{tipo_str}:<br>{lbl_val}{str_unidad}</b>", 
                showarrow=False, yshift=40, font=dict(size=11)
            )
    return fig

# --- 3. ETL YUPANA (PROGRAMADO) ---
archivos_clave_yupana = {
    "TERMICA"      : "Termica - Despacho (MW)",
    "COMBUSTIBLE"  : "Termica - Consumo de Combustible"
}

def clasificar_tecnologia_yupana(nombre_central):
    nombre = str(nombre_central).upper()
    diesel_kws = ["D2", "R6", "RESIDUAL", "DIESEL", "DIÉSEL", "ILO21", "ILO 21", "ILO1", "ILO 1", "MOLLENDO", "RECKA", "INDEPENDENCIA", "SAMANCO", "TARAPOTO", "IQUITOS", "YURIMAGUAS", "PUERTO MALDONADO", "BELLAVISTA", "PEDRO RUIZ", "ETEN", "PIURA D", "CALANA", "ELOR", "SHCUMMINS", "SNTV", "NEPI", "PUERTO BRAVO", "NODO"]
    if any(kw in nombre for kw in diesel_kws): return "Residual+Diésel D2"
    return "Otra Tecnología"

def cargar_df_desde_zip(zf, stem):
    for info in zf.infolist():
        nombre_base = info.filename.split('/')[-1]
        if stem in nombre_base and not nombre_base.startswith("~"):
            with zf.open(info) as f:
                if nombre_base.upper().endswith('.CSV'): 
                    try: return pd.read_csv(f, sep=None, engine='python')
                    except:
                        f.seek(0)
                        return pd.read_csv(f, sep=',')
    return None

def extraer_todas_centrales(df):
    series = {}
    if df is None or df.empty: return series
    invalid_cols = ["HORA", "TIEMPO", "FECHA", "ETAPA", "GENERADOR"]
    if df.shape[1] > 1:
        cols = [c for c in df.columns if not any(inv in str(c).upper() for inv in invalid_cols) and not str(c).startswith("Unnamed")]
        for c in cols: series[c] = pd.to_numeric(df[c], errors='coerce').fillna(0).tolist()
    else:
        enc = [h.strip() for h in str(df.columns[0]).split(",")]
        start_idx = 0
        if len(enc) < 2:
            enc = [h.strip() for h in str(df.iloc[0,0]).split(",")]
            start_idx = 1
        nombres_validos, idx_validos = [], []
        for i, nombre in enumerate(enc[1:], start=1):
            if not any(inv in nombre.upper() for inv in invalid_cols):
                nombres_validos.append(nombre)
                idx_validos.append(i)
                series[nombre] = []
        for fila in df.iloc[start_idx:, 0].astype(str):
            partes = [p.strip() for p in fila.split(",")]
            for nombre, i in zip(nombres_validos, idx_validos):
                series[nombre].append(float(partes[i]) if i < len(partes) and partes[i] else 0.0)
    return series

def extraer_motivo_dinamico(y, m, M, d, ddmm, l, headers):
    urls = [
        f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{l}%2FReprog_{ddmm}{l}.xlsx",
        f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog_{ddmm}{l}%2FReprog_{ddmm}{l}.xlsx"
    ]
    for u in urls:
        try:
            r = requests.get(u, headers=headers, timeout=10)
            if r.status_code == 200 and len(r.content) > 1000:
                wb = openpyxl.load_workbook(io.BytesIO(r.content), data_only=True)
                ws = wb.worksheets[0]
                for row in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=3).value
                    if cell_value and "MOTIVO" in str(cell_value).upper():
                        motivo_val = ws.cell(row=row+1, column=4).value
                        if motivo_val: return str(motivo_val).strip()
                return "No se encontró motivo."
        except: pass
    return "No se pudo extraer origen."

def rellenar_hasta_48(lst):
    if not lst: return [0.0]*48
    faltan = 48 - len(lst)
    return ([0.0]*faltan + lst) if faltan > 0 else lst[:48]

def suma_elementos_variable(*listas):
    if not listas: return []
    length = max(len(l) for l in listas if l)
    if length == 0: return []
    out = [0.0]*length
    for lst in listas:
        if lst:
            for i in range(min(length, len(lst))):
                if pd.notna(lst[i]): out[i] += lst[i]
    return out

@st.cache_data(show_spinner=False, ttl=300)
def extraer_datos_yupana_memoria(f):
    y, m, d = f.strftime("%Y"), f.strftime("%m"), f.strftime("%d")
    M = MES_TXT[f.month-1]
    fecha_str = f"{y}{m}{d}"
    ddmm = f"{d}{m}"
    
    headers = {'User-Agent': 'Mozilla/5.0'}
    datos_dia = {"Dataframes": {}, "Log": [], "stop_yupana": False}
    
    url_pdo = f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FPrograma%20Diario%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FYUPANA_{fecha_str}.zip"
    try:
        r = requests.get(url_pdo, headers=headers, timeout=15)
        if r.status_code == 200 and r.content[:4] == b'PK\x03\x04':
            with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
                datos_dia["Dataframes"]["PDO"] = {}
                for key, stem in archivos_clave_yupana.items():
                    datos_dia["Dataframes"]["PDO"][key] = extraer_todas_centrales(cargar_df_desde_zip(zf, stem))
            datos_dia["Log"].append("✅ PDO")
        else: 
            datos_dia["Log"].append("❌ PDO (Fallo - Deteniendo YUPANA)")
            datos_dia["stop_yupana"] = True
            return datos_dia
    except Exception: 
        datos_dia["Log"].append("❌ PDO (Error - Deteniendo YUPANA)")
        datos_dia["stop_yupana"] = True
        return datos_dia

    letra_actual = 'A'
    while True:
        nombre_rdo = f"RDO_{letra_actual}"
        urls_rdo = [
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}{letra_actual}%2FYUPANA_{ddmm}{letra_actual}.zip",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog_{ddmm}{letra_actual}%2FYUPANA_{ddmm}{letra_actual}.zip",
            f"https://www.coes.org.pe/portal/browser/download?url=Operaci%C3%B3n%2FPrograma%20de%20Operaci%C3%B3n%2FReprograma%20Diario%20Operaci%C3%B3n%2F{y}%2F{m}_{M}%2FD%C3%ADa%20{d}%2FReprog%20{ddmm}%20{letra_actual}%2FYUPANA_{ddmm}{letra_actual}.zip"
        ]
        
        exito_rdo = False
        for enlace in urls_rdo:
            if exito_rdo: break
            try:
                r = requests.get(enlace, headers=headers, timeout=10)
                if r.status_code == 200 and r.content[:4] == b'PK\x03\x04':
                    with zipfile.ZipFile(io.BytesIO(r.content)) as zf:
                        datos_dia["Dataframes"][nombre_rdo] = {}
                        for key, stem in archivos_clave_yupana.items():
                            datos_dia["Dataframes"][nombre_rdo][key] = extraer_todas_centrales(cargar_df_desde_zip(zf, stem))
                    datos_dia["Dataframes"][f"MOTIVO_{nombre_rdo}"] = extraer_motivo_dinamico(y, m, M, d, ddmm, letra_actual, headers)
                    datos_dia["Log"].append(f"✅ {nombre_rdo}")
                    exito_rdo = True
            except Exception: continue 
        
        if exito_rdo: letra_actual = chr(ord(letra_actual) + 1)
        else:
            datos_dia["Log"].append(f"🛑 Fin en {nombre_rdo}")
            break
            
    return datos_dia

def crear_grafica_area_apilada(df_plot, marcadores=None, unidad="Galones"):
    df_plot = df_plot.copy().fillna(0)
    num_cols = [c for c in df_plot.columns if c != 'Hora']
    df_plot[num_cols] = df_plot[num_cols].apply(pd.to_numeric, errors='coerce').fillna(0).round(2)
    
    df_plot['TOTAL_SISTEMA'] = df_plot[num_cols].sum(axis=1).round(2)
    totales_por_unidad = df_plot.drop(columns=['Hora', 'TOTAL_SISTEMA']).sum()
    orden_columnas = totales_por_unidad.sort_values(ascending=False).index.tolist()
    
    cols_mantener = ['Hora'] + orden_columnas
    df_melt = df_plot[cols_mantener].melt(id_vars=['Hora'], var_name='Unidad Generadora', value_name='Consumo')
    
    fig = px.area(df_melt, x="Hora", y="Consumo", color="Unidad Generadora", labels={"Consumo": "Consumo Físico"})
    fig.update_xaxes(tickformat="%d/%m %H:%M", tickangle=45)
    
    fig.add_scatter(x=df_plot['Hora'], y=df_plot['TOTAL_SISTEMA'], mode='lines', line=dict(width=0, color='rgba(0,0,0,0)'), name='<b>⚡ TOTAL CONSUMO</b>', showlegend=False)
    
    for trace in fig.data:
        trace.hoverinfo = ['skip' if pd.isna(v) or float(v) <= 0.01 else 'all' for v in trace.y]
        if 'TOTAL CONSUMO' in trace.name: 
            trace.hovertemplate = f'<b>%{{y:,.2f}} {unidad}</b><br>%{{x|%d/%m %H:%M}}'
        else: 
            trace.hovertemplate = f"%{{y:,.2f}} {unidad}"
    
    if marcadores:
        for ts, texto in marcadores:
            fig.add_vline(x=ts, line_width=1.5, line_dash="dash", line_color="rgba(255,255,255,0.7)")
            align = "left" if ts.hour == 0 and ts.minute == 30 else "center"
            fig.add_annotation(x=ts, y=1.02, yref="paper", text=f"<b>{texto} {ts.strftime('%H:%M')}</b>", showarrow=False, font=dict(size=10, color="white"), bgcolor="#e74c3c", bordercolor="white", borderwidth=1, borderpad=3, textangle=-90, yanchor="bottom", xanchor=align)
            
    fig.update_layout(hovermode="x unified", height=600, margin=dict(t=120, b=50, l=60, r=50), yaxis_title=f"Consumo de Combustible ({unidad})")
    return fig

# --- 4. ETL IEOD (EJECUTADO / POST-OPERACIÓN / STOCK) ---
def generar_urls_coes(fecha):
    año = fecha.strftime("%Y")
    mes_num = fecha.strftime("%m")
    dia = fecha.strftime("%d")
    mes_titulo = MESES_IEOD[fecha.month]
    fecha_str = fecha.strftime("%d%m")
    
    path_nuevo = f"Post Operación/Reportes/IEOD/{año}/{mes_num}_{mes_titulo}/{dia}/AnexoA_{fecha_str}.xlsx"
    path_legacy = f"Post Operación/Reportes/IEOD/{año}/{mes_num}_{mes_titulo}/{dia}/Anexo1_Resumen_{fecha_str}.xlsx"
    
    # ⚡ FIX TÉCNICO: Forzar safe='' para convertir los '/' en '%2F' (Requisito normativo del portal COES)
    return [
        (f"https://www.coes.org.pe/portal/browser/download?url={urllib.parse.quote(path_nuevo, safe='')}", "AnexoA"),
        (f"https://www.coes.org.pe/portal/browser/download?url={urllib.parse.quote(path_legacy, safe='')}", "Anexo1")
    ]

@st.cache_data(show_spinner=False)
def extraer_datos_ieod(fecha):
    urls = generar_urls_coes(fecha)
    headers = {'User-Agent': 'Mozilla/5.0'}
    df_raw = None
    df_stock = None
    
    for url, tipo_anexo in urls:
        try:
            res = requests.get(url, headers=headers, timeout=20)
            if res.status_code == 200:
                archivo_excel = io.BytesIO(res.content)
                xls = pd.ExcelFile(archivo_excel, engine='openpyxl')
                hojas_limpias = {h.strip().upper(): h for h in xls.sheet_names}
                
                if "CONSUMO_COMB" in hojas_limpias:
                    df_raw = pd.read_excel(xls, sheet_name=hojas_limpias["CONSUMO_COMB"], header=6, usecols="B:G")
                if "STOCK_COMB" in hojas_limpias:
                    df_stock = pd.read_excel(xls, sheet_name=hojas_limpias["STOCK_COMB"], header=5, usecols="C:E, G:I")
                
                if df_raw is not None or df_stock is not None: break
        except Exception: continue
            
    if (df_raw is None or df_raw.empty) and (df_stock is None or df_stock.empty): 
        return pd.DataFrame(), pd.DataFrame(), f"[{fecha.strftime('%d/%m/%Y')}] No se halló IEOD."

    if df_raw is not None and not df_raw.empty:
        try: df_raw.columns = ['EMPRESA', 'CENTRAL', 'MEDIDOR', 'TIPO_COMBUSTIBLE', 'UNIDAD_MEDIDA', 'CONSUMO']
        except ValueError: return pd.DataFrame(), pd.DataFrame(), f"[{fecha.strftime('%d/%m/%Y')}] Error estructura COES Consumo."
    
        df_raw = df_raw.dropna(subset=['EMPRESA', 'CENTRAL'])
        df_raw['CONSUMO'] = df_raw['CONSUMO'].astype(str).str.replace(',', '', regex=False)
        df_raw['CONSUMO'] = pd.to_numeric(df_raw['CONSUMO'], errors='coerce').fillna(0)
        
        for col in ['EMPRESA', 'CENTRAL', 'MEDIDOR', 'TIPO_COMBUSTIBLE', 'UNIDAD_MEDIDA']:
            df_raw[col] = df_raw[col].astype(str).str.strip().str.upper()

        df_raw = df_raw[~df_raw['TIPO_COMBUSTIBLE'].str.contains('GAS', na=False)]
        df_raw = df_raw[df_raw['TIPO_COMBUSTIBLE'].str.contains('DIESEL|DIÉSEL|RESIDUAL', na=False)]
        
        mask_diesel = df_raw['TIPO_COMBUSTIBLE'].str.contains('DIESEL|DIÉSEL|RESIDUAL', na=False)
        df_raw.loc[mask_diesel, 'UNIDAD_MEDIDA'] = 'M3'
        df_raw['FECHA_OPERATIVA'] = pd.to_datetime(fecha)
    else: df_raw = pd.DataFrame()

    if df_stock is not None and not df_stock.empty:
        try: df_stock.columns = ['EMPRESA', 'CENTRAL', 'TIPO_COMBUSTIBLE', 'STOCK_FINAL', 'REPOSICION', 'UNIDADES']
        except ValueError: return df_raw, pd.DataFrame(), f"[{fecha.strftime('%d/%m/%Y')}] Error estructura COES Stock."
        
        df_stock = df_stock.dropna(subset=['EMPRESA', 'CENTRAL'])
        for col in ['STOCK_FINAL', 'REPOSICION']:
            df_stock[col] = df_stock[col].astype(str).str.replace(',', '', regex=False)
            df_stock[col] = pd.to_numeric(df_stock[col], errors='coerce').fillna(0)
            
        for col in ['EMPRESA', 'CENTRAL', 'TIPO_COMBUSTIBLE']:
            df_stock[col] = df_stock[col].astype(str).str.strip().str.upper()
            
        df_stock = df_stock[~df_stock['TIPO_COMBUSTIBLE'].str.contains('GAS', na=False)]
        df_stock = df_stock[df_stock['TIPO_COMBUSTIBLE'].str.contains('DIESEL|DIÉSEL|RESIDUAL', na=False)]
        df_stock['UNIDADES'] = 'M3'
        df_stock['FECHA_OPERATIVA'] = pd.to_datetime(fecha)
    else: df_stock = pd.DataFrame()

    return df_raw, df_stock, None

# --- 5. INTERFAZ Y EJECUCIÓN ---
st.sidebar.header("Parámetros de Fiscalización")
rango_fechas = st.sidebar.date_input("Intervalo de Fechas", value=(date.today() - timedelta(days=2), date.today() - timedelta(days=1)))

st.sidebar.markdown("### Acciones de Extracción")
btn_extraer = st.sidebar.button("⚡ Extraer Datos (YUPANA e IEOD)", type="primary", use_container_width=True)

if btn_extraer:
    if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
        ini, fin = rango_fechas
        status, prog_bar = st.empty(), st.progress(0)
        
        log_exp = st.expander("Ver bitácora de descargas YUPANA", expanded=False)
        datos_completos_yupana = {}
        dias = (fin - ini).days + 1
        
        for k in range(dias):
            f_actual = ini + timedelta(days=k)
            status.markdown(f"**⏳ [1/2] Vectorizando YUPANA (Programado):** {f_actual.strftime('%d/%m/%Y')}")
            datos_dia = extraer_datos_yupana_memoria(f_actual)
            
            datos_completos_yupana[f_actual] = datos_dia
            with log_exp: st.markdown(f"**{f_actual.strftime('%d/%m/%Y')}** ➔ " + " | ".join(datos_dia["Log"]))
            prog_bar.progress(((k + 1) / dias) * 0.5)
            
            if datos_dia.get("stop_yupana"):
                with log_exp: st.markdown("⚠️ **Se detuvo la extracción de días posteriores en YUPANA debido a la falta del PDO base.**")
                break
            
        st.session_state['datos_yupana'] = datos_completos_yupana
        
        fechas_ieod = pd.date_range(ini, fin)
        total_dias_ieod = len(fechas_ieod)
        lista_dfs_ieod, lista_dfs_stock, alertas_ieod = [], [], []
        
        for i, f in enumerate(fechas_ieod):
            status.markdown(f"**⏳ [2/2] Procesando IEOD (Ejecutado e Inventario):** {f.strftime('%d/%m/%Y')}")
            df_dia, df_stk, error = extraer_datos_ieod(f)
            
            if not df_dia.empty: lista_dfs_ieod.append(df_dia)
            if not df_stk.empty: lista_dfs_stock.append(df_stk)
            if error: alertas_ieod.append(error)
            
            prog_bar.progress(0.5 + (((i + 1) / total_dias_ieod) * 0.5))
                
        if lista_dfs_ieod: st.session_state['df_ieod'] = pd.concat(lista_dfs_ieod, ignore_index=True)
        else: st.session_state['df_ieod'] = pd.DataFrame()
        
        if lista_dfs_stock: st.session_state['df_stock'] = pd.concat(lista_dfs_stock, ignore_index=True)
        else: st.session_state['df_stock'] = pd.DataFrame()
        
        st.session_state['alertas_ieod'] = alertas_ieod
        st.session_state['rango_extraccion'] = (ini, fin)
        
        status.success("✅ Motores YUPANA e IEOD Compilados con Éxito.")
        prog_bar.empty()

st.markdown("---")

# --- 6. VISUALIZACIÓN MULTI-PESTAÑA ---
t_yupana, t_ieod, t_proy = st.tabs([
    "📅 YUPANA (Programación y Motivos)", 
    "✅ IEOD Integral (Consumo, Stock y Reposición)",
    "🔮 Proyección Diaria (Estimación)"
])

# =====================================================================
# ====== TAB 1: YUPANA (PROGRAMADO Y MOTIVOS) ======
# =====================================================================
with t_yupana:
    st.info("**Contexto Osinergmin:** Datos del **Despacho Programado** de la Reserva Fría / Diésel, acompañados de sus **Justificaciones (Motivos RDO)**.")
    
    c_unidad_yup, _ = st.columns([1, 1])
    with c_unidad_yup:
        unidad_sel_yupana = st.radio("⚙️ Selección Volumétrica (YUPANA):", ["m3", "Galones", "bbl"], horizontal=True, key="rad_yupana")
    
    # Factor de conversión considerando que la data cruda de YUPANA está en Galones.
    factor_yupana = 1.0
    if unidad_sel_yupana == "m3": factor_yupana = 1 / 264.172
    elif unidad_sel_yupana == "bbl": factor_yupana = 1 / 42.0
    
    if 'datos_yupana' in st.session_state:
        data = st.session_state['datos_yupana']
        fechas_ordenadas = sorted(data.keys())
        active_prog_dict, ts_dict, dics_cache_dict, marcadores_globales = {}, {}, {}, []
        
        for f in fechas_ordenadas:
            df_dia_sel = data[f]["Dataframes"]
            progs = ["PDO"] if "PDO" in df_dia_sel else []
            progs.extend(sorted([p for p in df_dia_sel.keys() if p.startswith("RDO_")]))
            if not progs: continue
            
            dics_cache = {p: df_dia_sel[p] for p in progs}
            active_prog = [progs[0]] * 48
            if len(progs) > 1:
                for p in progs[1:]:
                    tot = [0.0]*48
                    for v in dics_cache[p].get("TERMICA", {}).values(): tot = suma_elementos_variable(tot, rellenar_hasta_48(v))
                    for i, val in enumerate(tot):
                        if val > 50:
                            for j in range(i, 48): active_prog[j] = p
                            break
                            
            ts_dia = [datetime.combine(f, datetime.min.time()) + timedelta(minutes=30*(i+1)) for i in range(48)]
            p_actual = active_prog[0]
            marcadores_globales.append((ts_dia[0], p_actual))
            for i in range(1, 48):
                if active_prog[i] != p_actual:
                    p_actual = active_prog[i]
                    marcadores_globales.append((ts_dia[i], p_actual))
                    
            active_prog_dict[f], ts_dict[f], dics_cache_dict[f] = active_prog, ts_dia, dics_cache

        dfs_comb, dfs_term = [], []
        for f in fechas_ordenadas:
            if f not in active_prog_dict: continue
            active_prog, dics_cache = active_prog_dict[f], dics_cache_dict[f]
            dia_data_comb, dia_data_term = {}, {}
            
            for i in range(48):
                p = active_prog[i]
                if "COMBUSTIBLE" in dics_cache[p] and dics_cache[p]["COMBUSTIBLE"]:
                    for central, v_list in dics_cache[p]["COMBUSTIBLE"].items():
                        c_clean = str(central).strip()
                        if clasificar_tecnologia_yupana(c_clean) == "Residual+Diésel D2":
                            if c_clean not in dia_data_comb: dia_data_comb[c_clean] = [0.0] * 48
                            dia_data_comb[c_clean][i] += rellenar_hasta_48(v_list)[i]
                            
                if "TERMICA" in dics_cache[p] and dics_cache[p]["TERMICA"]:
                    for central, v_list in dics_cache[p]["TERMICA"].items():
                        c_clean = str(central).strip()
                        if clasificar_tecnologia_yupana(c_clean) == "Residual+Diésel D2":
                            if c_clean not in dia_data_term: dia_data_term[c_clean] = [0.0] * 48
                            dia_data_term[c_clean][i] += rellenar_hasta_48(v_list)[i]
                            
            df_dia_comb = pd.DataFrame(dia_data_comb)
            df_dia_comb.insert(0, 'Hora', ts_dict[f])
            dfs_comb.append(df_dia_comb)
            
            df_dia_term = pd.DataFrame(dia_data_term)
            df_dia_term.insert(0, 'Hora', ts_dict[f])
            dfs_term.append(df_dia_term)
            
        if dfs_comb:
            df_total_comb = pd.concat(dfs_comb, ignore_index=True).fillna(0)
            df_total_term = pd.concat(dfs_term, ignore_index=True).fillna(0)
            
            num_cols = [c for c in df_total_comb.columns if c != 'Hora']
            active_cols = [c for c in num_cols if df_total_comb[c].sum() > 0]
            
            if not active_cols:
                st.success("✅ No hubo consumo PROGRAMADO de Diésel/Reserva Fría en YUPANA.")
            else:
                todas_centrales_comb = sorted(active_cols)
                filtro_comb = st.multiselect("⚙️ Filtrar Nodos YUPANA:", options=todas_centrales_comb, default=[])
                lista_filtro_comb = filtro_comb if filtro_comb else todas_centrales_comb
                
                df_daily = df_total_comb.copy()
                df_daily['Fecha_Operativa'] = (df_daily['Hora'] - pd.Timedelta(minutes=1)).dt.date
                df_daily_grouped = df_daily.groupby('Fecha_Operativa')[lista_filtro_comb].sum() * factor_yupana
                
                daily_totals = df_daily_grouped.sum(axis=1)
                max_centrals = df_daily_grouped.idxmax(axis=1)
                max_vals = df_daily_grouped.max(axis=1)
                
                st.markdown(f"#### 📊 Consumo Total Diario ({unidad_sel_yupana})")
                metric_cols = st.columns(len(daily_totals))
                for idx, (fecha_val, total_val) in enumerate(daily_totals.items()):
                    with metric_cols[idx]:
                        st.metric(
                            label=f"📅 {fecha_val.strftime('%d/%m/%Y')} (Prog.)", 
                            value=f"{formato_k_m(total_val)} {unidad_sel_yupana}",
                            delta=f"🔥 Max: {max_centrals[fecha_val]} ({formato_k_m(max_vals[fecha_val])} {unidad_sel_yupana})",
                            delta_color="off"
                        )
                st.markdown("---")
                
                df_plot_comb = df_total_comb[['Hora'] + lista_filtro_comb].copy()
                for c in lista_filtro_comb:
                    df_plot_comb[c] = df_plot_comb[c] * factor_yupana
                    
                st.plotly_chart(crear_grafica_area_apilada(df_plot_comb, marcadores=marcadores_globales, unidad=unidad_sel_yupana), use_container_width=True)
                
                st.markdown("#### 📋 Resumen Operativo Diario por Central (YUPANA)")
                df_c_melt = df_plot_comb.melt(id_vars=['Hora'], value_vars=lista_filtro_comb, var_name='Central', value_name='Consumo_Volumen')
                df_c_melt['Fecha'] = (df_c_melt['Hora'] - pd.Timedelta(minutes=1)).dt.strftime('%d/%m/%Y')
                res_comb = df_c_melt.groupby(['Fecha', 'Central'])['Consumo_Volumen'].sum().reset_index()
                
                cols_term = [c for c in lista_filtro_comb if c in df_total_term.columns]
                if cols_term:
                    df_t_melt = df_total_term.melt(id_vars=['Hora'], value_vars=cols_term, var_name='Central', value_name='Potencia_MW')
                    df_t_melt['Fecha'] = (df_t_melt['Hora'] - pd.Timedelta(minutes=1)).dt.strftime('%d/%m/%Y')
                    
                    df_t_activa = df_t_melt[df_t_melt['Potencia_MW'] > 0.01].copy()
                    
                    res_term_horas = df_t_activa.groupby(['Fecha', 'Central'])['Potencia_MW'].count().reset_index()
                    res_term_horas['Potencia_MW'] = res_term_horas['Potencia_MW'] * 0.5
                    res_term_horas.rename(columns={'Potencia_MW': 'Horas_Operacion'}, inplace=True)
                    
                    res_term_prom = df_t_activa.groupby(['Fecha', 'Central'])['Potencia_MW'].mean().reset_index()
                    res_term_prom.rename(columns={'Potencia_MW': 'Potencia_Promedio_MW'}, inplace=True)
                    
                    res_term = pd.merge(res_term_horas, res_term_prom, on=['Fecha', 'Central'], how='outer')
                    df_resumen = pd.merge(res_comb, res_term, on=['Fecha', 'Central'], how='left').fillna(0)
                else:
                    df_resumen = res_comb.copy()
                    df_resumen['Horas_Operacion'] = 0.0
                    df_resumen['Potencia_Promedio_MW'] = 0.0
                
                df_resumen = df_resumen[(df_resumen['Consumo_Volumen'] > 0) | (df_resumen['Horas_Operacion'] > 0)]
                
                if not df_resumen.empty:
                    df_view_yup = df_resumen.copy()
                    df_view_yup['Consumo_Volumen'] = df_view_yup['Consumo_Volumen'].apply(formato_k_m)
                    
                    st.dataframe(
                        df_view_yup, use_container_width=True, hide_index=True,
                        column_config={
                            "Fecha": "Día Operativo", "Central": "Central Térmica",
                            "Consumo_Volumen": f"Consumo Programado ({unidad_sel_yupana})",
                            "Horas_Operacion": st.column_config.NumberColumn("Horas de Operación", format="%.1f h"),
                            "Potencia_Promedio_MW": st.column_config.NumberColumn("Potencia Promedio", format="%.2f MW")
                        }
                    )
                    st.download_button(
                        label="📥 Descargar Resumen YUPANA (CSV)",
                        data=df_resumen.to_csv(index=False).encode('utf-8'),
                        file_name="resumen_yupana.csv",
                        mime="text/csv",
                    )
                else: st.info("No hay operación registrada para las centrales seleccionadas.")
                
        st.markdown("---")
        st.markdown("#### 📋 Motivos de Reprogramación Operativa (RDO)")
        tabla_motivos = []
        for f in fechas_ordenadas:
            if f not in dics_cache_dict: continue
            for p in sorted(set(active_prog_dict[f])):
                if "RDO" in p:
                    tabla_motivos.append({
                        "Fecha": f.strftime("%d/%m/%Y"), 
                        "Reprograma": p, 
                        "Motivo Declarado": st.session_state['datos_yupana'][f]["Dataframes"].get(f"MOTIVO_{p}", "No disponible.")
                    })
        if tabla_motivos: st.dataframe(pd.DataFrame(tabla_motivos), use_container_width=True)
        else: st.success("Sin reprogramas justificados en el periodo extraído.")

    else: st.warning("👈 Por favor, configura las fechas y haz clic en **'⚡ Extraer Datos'**.")


# =====================================================================
# ====== TAB 2: IEOD INTEGRAL (Puro y Real - Barras Agrupadas) ======
# =====================================================================
with t_ieod:
    st.info("**Contexto Osinergmin:** Entorno de auditoría pura. Muestra únicamente **Datos Reales Ejecutados** (IEOD). Las gráficas se muestran en formato de barras agrupadas y apiladas para facilitar distintas comparativas entre centrales de un mismo día.")

    has_stock = 'df_stock' in st.session_state and not st.session_state['df_stock'].empty
    has_ieod = 'df_ieod' in st.session_state and not st.session_state['df_ieod'].empty

    if has_stock or has_ieod:
        df_stock_log = st.session_state['df_stock'] if has_stock else pd.DataFrame()
        df_ieod_log = st.session_state['df_ieod'] if has_ieod else pd.DataFrame()
        
        c_unidad_log, _ = st.columns([1, 1])
        with c_unidad_log:
            unidad_sel_log = st.radio("⚙️ Selección Volumétrica Integral (Solo IEOD):", ["m3", "Galones", "bbl"], horizontal=True, key="rad_ieod_puro")

        empresas_totales = set()
        if not df_stock_log.empty: empresas_totales.update(df_stock_log['EMPRESA'].dropna().unique())
        if not df_ieod_log.empty: empresas_totales.update(df_ieod_log['EMPRESA'].dropna().unique())
        
        st.markdown("#### 🔍 Filtros Transversales")
        c1, c2, c3 = st.columns(3)
        with c1: filtro_emp_ieod = st.multiselect("🏢 Empresa:", options=sorted(list(empresas_totales)), key="emp_ieod_puro")
        
        centrales_totales = set()
        if not df_stock_log.empty: centrales_totales.update(df_stock_log[df_stock_log['EMPRESA'].isin(filtro_emp_ieod)]['CENTRAL'].unique() if filtro_emp_ieod else df_stock_log['CENTRAL'].unique())
        if not df_ieod_log.empty: centrales_totales.update(df_ieod_log[df_ieod_log['EMPRESA'].isin(filtro_emp_ieod)]['CENTRAL'].unique() if filtro_emp_ieod else df_ieod_log['CENTRAL'].unique())
        
        with c2: filtro_cen_ieod = st.multiselect("⚡ Central:", options=sorted(list(centrales_totales)), key="cen_ieod_puro")
        
        combs_totales = set()
        if not df_stock_log.empty: combs_totales.update(df_stock_log[df_stock_log['CENTRAL'].isin(filtro_cen_ieod)]['TIPO_COMBUSTIBLE'].unique() if filtro_cen_ieod else df_stock_log['TIPO_COMBUSTIBLE'].unique())
        if not df_ieod_log.empty: combs_totales.update(df_ieod_log[df_ieod_log['CENTRAL'].isin(filtro_cen_ieod)]['TIPO_COMBUSTIBLE'].unique() if filtro_cen_ieod else df_ieod_log['TIPO_COMBUSTIBLE'].unique())
        
        with c3: filtro_comb_ieod = st.multiselect("🛢️ Combustible:", options=sorted(list(combs_totales)), key="comb_ieod_puro")

        if not df_stock_log.empty:
            if filtro_emp_ieod: df_stock_log = df_stock_log[df_stock_log['EMPRESA'].isin(filtro_emp_ieod)]
            if filtro_cen_ieod: df_stock_log = df_stock_log[df_stock_log['CENTRAL'].isin(filtro_cen_ieod)]
            if filtro_comb_ieod: df_stock_log = df_stock_log[df_stock_log['TIPO_COMBUSTIBLE'].isin(filtro_comb_ieod)]
            
        if not df_ieod_log.empty:
            if filtro_emp_ieod: df_ieod_log = df_ieod_log[df_ieod_log['EMPRESA'].isin(filtro_emp_ieod)]
            if filtro_cen_ieod: df_ieod_log = df_ieod_log[df_ieod_log['CENTRAL'].isin(filtro_cen_ieod)]
            if filtro_comb_ieod: df_ieod_log = df_ieod_log[df_ieod_log['TIPO_COMBUSTIBLE'].isin(filtro_comb_ieod)]

        comb_iterar = filtro_comb_ieod if filtro_comb_ieod else list(combs_totales)

        for comb in sorted(comb_iterar):
            st.markdown("---")
            st.markdown(f"### 🛢️ Balance Logístico: {comb} (IEOD Real)")
            
            if not df_ieod_log.empty:
                df_p_cons = df_ieod_log[df_ieod_log['TIPO_COMBUSTIBLE'] == comb].copy()
                if not df_p_cons.empty:
                    df_p_cons['CONS_PLOT'] = convertir_volumen(df_p_cons['CONSUMO'], df_p_cons['UNIDAD_MEDIDA'], unidad_sel_log)
                    df_cons_final = df_p_cons[df_p_cons['CONS_PLOT'] > 0].groupby(['FECHA_OPERATIVA', 'CENTRAL'])['CONS_PLOT'].sum().reset_index()
                    df_cons_final = df_cons_final.sort_values(by=['FECHA_OPERATIVA', 'CENTRAL'])
                    
                    if not df_cons_final.empty:
                        st.markdown(f"### 🔥 Consumo Diario Ejecutado ({unidad_sel_log})")
                        tab_cons_agrupada, tab_cons_apilada = st.tabs(["📊 Vista Agrupada", "📊 Vista Apilada"])
                        
                        with tab_cons_agrupada:
                            # Gráfica de Barras AGRUPADAS
                            fig_cons_grp = px.bar(
                                df_cons_final, x="FECHA_OPERATIVA", y="CONS_PLOT", color="CENTRAL", text_auto=',.1f'
                            )
                            fig_cons_grp.update_layout(height=450, xaxis_title="Día Operativo", yaxis_title=f"Consumo ({unidad_sel_log})", barmode="group")
                            fig_cons_grp.update_traces(textposition='outside')
                            fig_cons_grp.update_xaxes(type='date', dtick="86400000", tickformat="%d/%m/%Y")
                            fig_cons_grp = agregar_totales_diarios(fig_cons_grp, df_cons_final, "FECHA_OPERATIVA", "CONS_PLOT", "Consumo Total", unidad_sel_log, barmode="group")
                            st.plotly_chart(fig_cons_grp, use_container_width=True)

                        with tab_cons_apilada:
                            # Gráfica de Barras APILADAS
                            fig_cons_stk = px.bar(
                                df_cons_final, x="FECHA_OPERATIVA", y="CONS_PLOT", color="CENTRAL", text_auto=',.1f'
                            )
                            fig_cons_stk.update_layout(height=450, xaxis_title="Día Operativo", yaxis_title=f"Consumo ({unidad_sel_log})", barmode="relative")
                            fig_cons_stk.update_traces(textposition='inside')
                            fig_cons_stk.update_xaxes(type='date', dtick="86400000", tickformat="%d/%m/%Y")
                            fig_cons_stk = agregar_totales_diarios(fig_cons_stk, df_cons_final, "FECHA_OPERATIVA", "CONS_PLOT", "Consumo Total", unidad_sel_log, barmode="relative")
                            st.plotly_chart(fig_cons_stk, use_container_width=True)

            if not df_stock_log.empty:
                df_p_stk = df_stock_log[df_stock_log['TIPO_COMBUSTIBLE'] == comb].copy()
                if not df_p_stk.empty:
                    df_p_stk['STOCK_PLOT'] = convertir_volumen(df_p_stk['STOCK_FINAL'], df_p_stk['UNIDADES'], unidad_sel_log)
                    df_p_stk['REPO_PLOT'] = convertir_volumen(df_p_stk['REPOSICION'], df_p_stk['UNIDADES'], unidad_sel_log)
                    
                    df_stk_final = df_p_stk[df_p_stk['STOCK_PLOT'] > 0].copy()
                    df_stk_final = df_stk_final.sort_values(by=['FECHA_OPERATIVA', 'CENTRAL'])
                    if not df_stk_final.empty:
                        st.markdown(f"### 🏭 Inventario Cierre del Día ({unidad_sel_log})")
                        tab_stk_agrupada, tab_stk_apilada = st.tabs(["📊 Vista Agrupada", "📊 Vista Apilada"])
                        
                        with tab_stk_agrupada:
                            # Gráfica de Barras AGRUPADAS
                            fig_stk_grp = px.bar(
                                df_stk_final, x="FECHA_OPERATIVA", y="STOCK_PLOT", color="CENTRAL", text_auto=',.1f'
                            )
                            fig_stk_grp.update_layout(height=450, xaxis_title="Día Operativo", yaxis_title=f"Stock Final ({unidad_sel_log})", barmode="group")
                            fig_stk_grp.update_traces(textposition='outside')
                            fig_stk_grp.update_xaxes(type='date', dtick="86400000", tickformat="%d/%m/%Y")
                            fig_stk_grp = agregar_totales_diarios(fig_stk_grp, df_stk_final, "FECHA_OPERATIVA", "STOCK_PLOT", "Stock Total", unidad_sel_log, barmode="group")
                            st.plotly_chart(fig_stk_grp, use_container_width=True)

                        with tab_stk_apilada:
                            # Gráfica de Barras APILADAS
                            fig_stk_stk = px.bar(
                                df_stk_final, x="FECHA_OPERATIVA", y="STOCK_PLOT", color="CENTRAL", text_auto=',.1f'
                            )
                            fig_stk_stk.update_layout(height=450, xaxis_title="Día Operativo", yaxis_title=f"Stock Final ({unidad_sel_log})", barmode="relative")
                            fig_stk_stk.update_traces(textposition='inside')
                            fig_stk_stk.update_xaxes(type='date', dtick="86400000", tickformat="%d/%m/%Y")
                            fig_stk_stk = agregar_totales_diarios(fig_stk_stk, df_stk_final, "FECHA_OPERATIVA", "STOCK_PLOT", "Stock Total", unidad_sel_log, barmode="relative")
                            st.plotly_chart(fig_stk_stk, use_container_width=True)
                            
                    df_repo_final = df_p_stk[df_p_stk['REPO_PLOT'] > 0].copy()
                    df_repo_final = df_repo_final.sort_values(by=['FECHA_OPERATIVA', 'CENTRAL'])
                    if not df_repo_final.empty:
                        st.markdown(f"### 🚛 Reposición Diaria ({unidad_sel_log})")
                        tab_repo_agrupada, tab_repo_apilada = st.tabs(["📊 Vista Agrupada", "📊 Vista Apilada"])
                        
                        with tab_repo_agrupada:
                            # Gráfica de Barras AGRUPADAS
                            fig_repo_grp = px.bar(
                                df_repo_final, x="FECHA_OPERATIVA", y="REPO_PLOT", color="CENTRAL", text_auto=',.1f'
                            )
                            fig_repo_grp.update_layout(height=450, xaxis_title="Día Operativo", yaxis_title=f"Reposición Diaria ({unidad_sel_log})", barmode="group")
                            fig_repo_grp.update_traces(textposition='outside')
                            fig_repo_grp.update_xaxes(type='date', dtick="86400000", tickformat="%d/%m/%Y")
                            fig_repo_grp = agregar_totales_diarios(fig_repo_grp, df_repo_final, "FECHA_OPERATIVA", "REPO_PLOT", "Reposición Total", unidad_sel_log, barmode="group")
                            st.plotly_chart(fig_repo_grp, use_container_width=True)

                        with tab_repo_apilada:
                            # Gráfica de Barras APILADAS
                            fig_repo_stk = px.bar(
                                df_repo_final, x="FECHA_OPERATIVA", y="REPO_PLOT", color="CENTRAL", text_auto=',.1f'
                            )
                            fig_repo_stk.update_layout(height=450, xaxis_title="Día Operativo", yaxis_title=f"Reposición Diaria ({unidad_sel_log})", barmode="relative")
                            fig_repo_stk.update_traces(textposition='inside')
                            fig_repo_stk.update_xaxes(type='date', dtick="86400000", tickformat="%d/%m/%Y")
                            fig_repo_stk = agregar_totales_diarios(fig_repo_stk, df_repo_final, "FECHA_OPERATIVA", "REPO_PLOT", "Reposición Total", unidad_sel_log, barmode="relative")
                            st.plotly_chart(fig_repo_stk, use_container_width=True)
                    else:
                        st.info(f"**🚛 Reposición Diaria:** No hay reposición registrada para {comb} en el periodo y centrales seleccionadas.")

                    # =====================================================================
                    # ====== NUEVO MÓDULO: AUTONOMÍA DE OPERACIÓN A POTENCIA EFECTIVA ======
                    # =====================================================================
                    if not df_stk_final.empty:
                        ultima_fecha_stk = df_stk_final['FECHA_OPERATIVA'].max()
                        fecha_corte_str = ultima_fecha_stk.strftime('%d/%m/%Y')
                        
                        st.markdown("---")
                        st.markdown(f"### ⏱️ Autonomía Operativa Crítica: {comb} (Desde el {fecha_corte_str})")
                        st.info("💡 **Metodología de Cálculo (Operación a Plena Carga):**\n"
                                "La autonomía operativa se evalúa bajo el escenario de máxima exigencia para garantizar la seguridad del SEIN. "
                                "Se determina tomando el **Inventario de Cierre de Día (Stock Final)** convertido a galones, "
                                "dividido entre el **Consumo Diario Proyectado operando a Potencia Efectiva continua (24 horas)**. "
                                "El consumo diario a plena carga se halla dividiendo la energía máxima teórica diaria "
                                "(Potencia Efectiva [MW] × 24h × 1000) entre el Rendimiento Térmico (kWh/gal) auditado de cada central.")
                        
                        datos_centrales = {
                            "CHILINA": {"rendimiento": 11.21175382, "potencia": 22.516},
                            "MOLLENDO": {"rendimiento": 15.64626147, "potencia": 24.462},
                            "MALACAS": {"rendimiento": 14.04159985, "potencia": 43.26},
                            "TALARA": {"rendimiento": 14.00602638, "potencia": 184.875},
                            "NEPI": {"rendimiento": 13.87997446, "potencia": 600.0},
                            "ILO": {"rendimiento": 13.86753981, "potencia": 460.0},
                            "FENIX": {"rendimiento": 18.92706267, "potencia": 545.23399},
                            "PTO MALDONADO": {"rendimiento": 12.49186136, "potencia": 17.416},
                            "MALDONADO": {"rendimiento": 12.49186136, "potencia": 17.416},
                            "PUCALLPA": {"rendimiento": 12.11332011, "potencia": 44.054},
                            "RECKA": {"rendimiento": 13.24894387, "potencia": 179.373},
                            "SANTA ROSA": {"rendimiento": 11.4447210, "potencia": 207.4278}, # Parámetros calibrados
                            "VENTANILLA": {"rendimiento": 13.24894387, "potencia": 279.720},
                            "ETEN": {"rendimiento": 13.77425577, "potencia": 225.055},
                            "PUERTO BRAVO": {"rendimiento": 13.62748512, "potencia": 600.0},
                            "SAN NICOLAS": {"rendimiento": 11.25381985, "potencia": 62.820},
                            "TUMBES": {"rendimiento": 17.79143891, "potencia": 17.343}
                        }

                        def obtener_parametros(nombre_central):
                            for clave, params in datos_centrales.items():
                                if clave in str(nombre_central).upper():
                                    return params['rendimiento'], params['potencia']
                            return 13.5, 100.0  # Valores referenciales por defecto si no se encuentra en el listado

                        df_autonomia = df_stk_final[df_stk_final['FECHA_OPERATIVA'] == ultima_fecha_stk].copy()
                        
                        # Conversión segura del stock a galones usando la función existente
                        df_autonomia['STOCK_GALONES'] = convertir_volumen(df_autonomia['STOCK_FINAL'], df_autonomia['UNIDADES'], 'Galones')
                        
                        # Obtener Parámetros Técnicos
                        df_autonomia[['RENDIMIENTO', 'POTENCIA_EFECTIVA']] = df_autonomia['CENTRAL'].apply(lambda x: pd.Series(obtener_parametros(x)))
                        
                        # Cálculo: Energía Máxima en 24h (kWh) y Consumo a Plena Carga (Galones)
                        df_autonomia['ENERGIA_DIA_KWH'] = df_autonomia['POTENCIA_EFECTIVA'] * 1000 * 24
                        df_autonomia['CONSUMO_PLENA_CARGA_GAL'] = df_autonomia['ENERGIA_DIA_KWH'] / df_autonomia['RENDIMIENTO']
                        
                        # Autonomía en días (se evita división por cero) a 2 decimales
                        df_autonomia['DIAS_AUTONOMIA'] = df_autonomia.apply(
                            lambda row: (row['STOCK_GALONES'] / row['CONSUMO_PLENA_CARGA_GAL']) if row['CONSUMO_PLENA_CARGA_GAL'] > 0 else 0, 
                            axis=1
                        ).round(2)

                        # Gráfica de Autonomía Similar al Consumo Diario Ejecutado
                        fig_auto = px.bar(
                            df_autonomia, 
                            x="CENTRAL", 
                            y="DIAS_AUTONOMIA", 
                            color="CENTRAL", 
                            text_auto='.2f',
                            custom_data=["POTENCIA_EFECTIVA", "RENDIMIENTO", "STOCK_PLOT"]
                        )
                        
                        fig_auto.update_layout(
                            height=450, 
                            xaxis_title="Central Térmica", 
                            yaxis_title="Autonomía Crítica (Días)",
                            showlegend=False,
                            barmode="group"
                        )
                        
                        fig_auto.update_traces(
                            textposition='outside',
                            hovertemplate=(
                                "<b>%{x}</b><br>"
                                "Autonomía a Plena Carga: %{y:.2f} Días<br>"
                                "Potencia Efectiva: %{customdata[0]:.4f} MW<br>"
                                "Rendimiento: %{customdata[1]:.7f} kWh/gal<br>"
                                "Stock Base Cierre: %{customdata[2]:,.2f} " + unidad_sel_log + "<extra></extra>"
                            )
                        )
                        
                        max_y = df_autonomia['DIAS_AUTONOMIA'].max()
                        fig_auto.update_layout(yaxis=dict(range=[0, max_y * 1.25 if max_y > 0 else 1]))
                        
                        st.plotly_chart(fig_auto, use_container_width=True)
                        
                        # Resumen tabular desplegable
                        with st.expander("📋 Ver Matriz de Autonomía Calculada a Potencia Efectiva"):
                            df_resumen_auto = df_autonomia[['CENTRAL', 'STOCK_PLOT', 'POTENCIA_EFECTIVA', 'RENDIMIENTO', 'CONSUMO_PLENA_CARGA_GAL', 'DIAS_AUTONOMIA']].copy()
                            df_resumen_auto.columns = ['Central', f'Stock de Cierre ({unidad_sel_log})', 'Potencia Efectiva (MW)', 'Rendimiento (kWh/gal)', 'Consumo a Plena Carga 24h (Galones)', 'Días de Autonomía']
                            st.dataframe(df_resumen_auto, use_container_width=True, hide_index=True)

        st.markdown("---")
        st.markdown("#### 🗄️ Trazabilidad de Registros Crudos Originales (Normativo en M3)")
        col_exp1, col_exp2 = st.columns(2)
        with col_exp1:
            with st.expander("Ver Datos Base de Consumo (IEOD)"):
                if not df_ieod_log.empty:
                    df_m_cons = df_ieod_log[['FECHA_OPERATIVA', 'EMPRESA', 'CENTRAL', 'MEDIDOR', 'TIPO_COMBUSTIBLE', 'UNIDAD_MEDIDA', 'CONSUMO']].copy()
                    df_m_cons['FECHA_OPERATIVA'] = df_m_cons['FECHA_OPERATIVA'].dt.strftime('%d/%m/%Y')
                    st.dataframe(df_m_cons, use_container_width=True, hide_index=True)
                else: st.write("Sin datos base de consumo IEOD.")
        with col_exp2:
            with st.expander("Ver Datos Base de Stock y Reposición (IEOD)"):
                if not df_stock_log.empty:
                    df_m_stk = df_stock_log[['FECHA_OPERATIVA', 'EMPRESA', 'CENTRAL', 'TIPO_COMBUSTIBLE', 'UNIDADES', 'STOCK_FINAL', 'REPOSICION']].copy()
                    df_m_stk['FECHA_OPERATIVA'] = df_m_stk['FECHA_OPERATIVA'].dt.strftime('%d/%m/%Y')
                    st.dataframe(df_m_stk, use_container_width=True, hide_index=True)
                else: st.write("Sin datos base de stock/reposición IEOD.")
    else:
        st.warning("👈 Por favor, configura las fechas y haz clic en **'⚡ Extraer Datos'** para visualizar el IEOD Integral.")


# =====================================================================
# ====== TAB 3: PROYECCIÓN DIARIA (ALGORITMO HOMÓLOGO DINÁMICO) ======
# =====================================================================
with t_proy:
    st.markdown("### 📋 Metodología de Proyección de Consumo")
    
    # --- NUEVOS CONTROLES DE PARAMETRIZACIÓN ---
    st.markdown("#### 🎛️ Parámetros de Reducción Semanal (Perfil de Despacho)")
    col_pct1, col_pct2, col_pct3 = st.columns(3)
    with col_pct1:
        pct_lunes = st.number_input("📉 Reducción Lunes (%)", min_value=0.0, max_value=100.0, value=5.0, step=0.5, format="%.1f")
    with col_pct2:
        pct_sabado = st.number_input("📉 Reducción Sábado (%)", min_value=0.0, max_value=100.0, value=7.5, step=0.5, format="%.1f")
    with col_pct3:
        pct_domingo = st.number_input("📉 Reducción Domingo (%)", min_value=0.0, max_value=100.0, value=23.0, step=0.5, format="%.1f")

    st.info(f"""
    **¿Cómo funciona el modelo de estimación para los días sin información?**
    1. **Línea Base Histórica:** El algoritmo toma como referencia estricta la información real (IEOD ejecutado) documentada a partir del **03 de Marzo de 2026** (excluyendo el 02 de marzo por ser un día atípico de alta demanda).
    2. **Estimación Base:** Se calcula el promedio de consumo de los últimos 2 días hábiles (Martes a Viernes) con información real disponible (IEOD) por cada central térmica.
    3. **Regla de Operación Semanal Dinámica:** - **Martes a Viernes:** Consumo estimado igual al promedio base calculado.
       - **Lunes:** Reducción del **{pct_lunes}%** respecto al promedio base.
       - **Sábados:** Reducción del **{pct_sabado}%** respecto al promedio base.
       - **Domingos:** Reducción del **{pct_domingo}%** respecto al promedio base.
    4. **Fusión Dinámica:** En el rango de fechas seleccionado, el gráfico muestra el dato real (**Sólido**) y aplica la proyección donde falten datos (**Achurado**).
    """)

    has_data_proy = ('df_ieod' in st.session_state and not st.session_state['df_ieod'].empty)

    if has_data_proy:
        df_ieod_proy = st.session_state['df_ieod'].copy()
        
        st.markdown("#### ⚙️ Filtros de Proyección")
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            ini_def = st.session_state['rango_extraccion'][0]
            fin_def = st.session_state['rango_extraccion'][1] + timedelta(days=7)
            rango_proy = st.date_input("📅 Rango de Fechas a Visualizar (Histórico + Proyección):", 
                                       value=(ini_def, fin_def), key="rango_input_proy")
        with col_f2:
            unidad_sel_proy = st.radio("Volumetría (Diésel/Residual):", ["m3", "Galones", "bbl"], horizontal=True, key="rad_proy_int_2")

        empresas_totales_proy = sorted(df_ieod_proy['EMPRESA'].dropna().unique()) if not df_ieod_proy.empty else []
        c1_p, c2_p, c3_p = st.columns(3)
        with c1_p: filtro_emp_proy = st.multiselect("🏢 Empresa:", options=empresas_totales_proy, key="emp_proy")
        
        df_ieod_f1 = df_ieod_proy[df_ieod_proy['EMPRESA'].isin(filtro_emp_proy)] if filtro_emp_proy else df_ieod_proy
        centrales_totales_proy = sorted(df_ieod_f1['CENTRAL'].dropna().unique()) if not df_ieod_f1.empty else []
        with c2_p: filtro_cen_proy = st.multiselect("⚡ Central:", options=centrales_totales_proy, key="cen_proy")
        
        df_ieod_f2 = df_ieod_f1[df_ieod_f1['CENTRAL'].isin(filtro_cen_proy)] if filtro_cen_proy else df_ieod_f1
        combs_totales_proy = sorted(df_ieod_f2['TIPO_COMBUSTIBLE'].dropna().unique()) if not df_ieod_f2.empty else []
        with c3_p: filtro_comb_proy = st.multiselect("🛢️ Combustible:", options=combs_totales_proy, key="comb_proy")

        st.markdown("---")

        comb_iterar_proy = filtro_comb_proy if filtro_comb_proy else combs_totales_proy
        
        for comb in comb_iterar_proy:
            df_ieod_liq = df_ieod_proy[df_ieod_proy['TIPO_COMBUSTIBLE'] == comb].copy()
            if filtro_emp_proy: df_ieod_liq = df_ieod_liq[df_ieod_liq['EMPRESA'].isin(filtro_emp_proy)]
            if filtro_cen_proy: df_ieod_liq = df_ieod_liq[df_ieod_liq['CENTRAL'].isin(filtro_cen_proy)]
            
            if df_ieod_liq.empty: continue
            
            st.markdown(f"### 🔮 Proyección Predictiva: {comb}")

            if isinstance(rango_proy, tuple) and len(rango_proy) == 2:
                df_ieod_liq['VAL_CONV'] = convertir_volumen(df_ieod_liq['CONSUMO'], df_ieod_liq['UNIDAD_MEDIDA'], unidad_sel_proy)
                
                # Línea base estricta desde el 03 de Marzo de 2026
                ref_date = pd.to_datetime("2026-03-03")
                df_base = df_ieod_liq[df_ieod_liq['FECHA_OPERATIVA'] >= ref_date].copy()
                if df_base.empty: df_base = df_ieod_liq 
                
                # --- CONSTRUCCIÓN DE REGLAS BASADAS EN PROMEDIO MARTES A VIERNES ---
                reglas_cen = {}
                for cen in df_base['CENTRAL'].unique():
                    df_cen = df_base[df_base['CENTRAL'] == cen]
                    df_cen_diario = df_cen.groupby('FECHA_OPERATIVA')['VAL_CONV'].sum().reset_index()
                    df_cen_diario = df_cen_diario.sort_values('FECHA_OPERATIVA')
                    
                    df_cen_diario['WEEKDAY'] = df_cen_diario['FECHA_OPERATIVA'].dt.weekday
                    df_martes_viernes = df_cen_diario[df_cen_diario['WEEKDAY'].isin([1, 2, 3, 4])]
                    ultimos_2_mv = df_martes_viernes.tail(2)
                    promedio_base = ultimos_2_mv['VAL_CONV'].mean() if not ultimos_2_mv.empty else 0.0
                    reglas_cen[cen] = promedio_base
                
                d_start, d_end = rango_proy
                dates_range = pd.date_range(d_start, d_end).date
                
                combined_records = []
                
                for d in dates_range:
                    d_ts = pd.to_datetime(d)
                    df_day = df_ieod_liq[df_ieod_liq['FECHA_OPERATIVA'].dt.date == d]
                    wd = d_ts.weekday() # 0 = Lunes, 1 = Martes, ..., 6 = Domingo
                    
                    if not df_day.empty:
                        day_grouped = df_day.groupby('CENTRAL')['VAL_CONV'].sum().reset_index()
                        for _, row in day_grouped.iterrows():
                            if row['VAL_CONV'] > 0:
                                combined_records.append({
                                    'FECHA_OPERATIVA': d_ts,
                                    'CENTRAL': row['CENTRAL'],
                                    'CONS_PLOT': row['VAL_CONV'],
                                    'TIPO_DATO': 'Ejecutado'
                                })
                    else:
                        # Estimación algorítmica perfilada con parámetros dinámicos del usuario
                        for cen, base_val in reglas_cen.items():
                            proj_val = 0.0
                            if wd in [1, 2, 3, 4]: # Martes a Viernes (Base regular)
                                proj_val = base_val
                            elif wd == 0: # Lunes
                                proj_val = base_val * (1 - (pct_lunes / 100.0))
                            elif wd == 5: # Sábado
                                proj_val = base_val * (1 - (pct_sabado / 100.0))
                            elif wd == 6: # Domingo
                                proj_val = base_val * (1 - (pct_domingo / 100.0))
                            
                            if proj_val > 0:
                                combined_records.append({
                                    'FECHA_OPERATIVA': d_ts,
                                    'CENTRAL': cen,
                                    'CONS_PLOT': proj_val,
                                    'TIPO_DATO': 'Proyectado'
                                })
                                
                df_combined = pd.DataFrame(combined_records)
                               
                if not df_combined.empty:
                    st.markdown(f"#### 📈 Evolución Diaria: Ejecutado vs. Proyectado")
                    
                    df_combined['DIA_SEMANA'] = df_combined['FECHA_OPERATIVA'].dt.weekday.map(lambda x: DIAS_ESP[x])
                    df_combined['FECHA_DISPLAY'] = df_combined['DIA_SEMANA'] + "<br>" + df_combined['FECHA_OPERATIVA'].dt.strftime("%d/%m/%Y")
                    
                    # Gráfico 1: Barras Apiladas por Día 
                    fig_proy = px.bar(
                        df_combined, x="FECHA_DISPLAY", y="CONS_PLOT", color="CENTRAL", pattern_shape="TIPO_DATO", text_auto=',.1f'
                    )
                    fig_proy.update_layout(height=500, xaxis_title="Día Operativo", yaxis_title=f"Consumo ({unidad_sel_proy})", barmode="relative")
                    fig_proy.update_xaxes(categoryorder='array', categoryarray=df_combined['FECHA_DISPLAY'].unique())
                    fig_proy = agregar_totales_diarios(fig_proy, df_combined, "FECHA_DISPLAY", "CONS_PLOT", "Total", unidad_sel_proy, col_tipo="TIPO_DATO", barmode="relative")
                    st.plotly_chart(fig_proy, use_container_width=True)

                    st.markdown("---")
                    
                    st.markdown(f"#### 📊 Consumo Total Acumulado por Central en el Periodo")
                    # Gráfico 2: Barras Apiladas por TIPO_DATO dentro de cada CENTRAL
                    df_total_cen = df_combined.groupby(['CENTRAL', 'TIPO_DATO'])['CONS_PLOT'].sum().reset_index()
                    
                    # Ordenamos las centrales por el gran total de mayor a menor
                    orden_centrales = df_total_cen.groupby('CENTRAL')['CONS_PLOT'].sum().sort_values(ascending=False).index.tolist()
                    
                    fig_tot = px.bar(
                        df_total_cen, x="CENTRAL", y="CONS_PLOT", color="CENTRAL", 
                        pattern_shape="TIPO_DATO", text_auto=',.1f',
                        category_orders={"CENTRAL": orden_centrales}
                    )
                    
                    fig_tot.update_layout(height=500, xaxis_title="Central Térmica", yaxis_title=f"Consumo Acumulado ({unidad_sel_proy})", barmode="relative")
                    
                    # Corregimos el cálculo del rango Y global para el gráfico de totales
                    df_tot_sum = df_total_cen.groupby('CENTRAL', as_index=False)['CONS_PLOT'].sum()
                    max_y_tot = df_tot_sum['CONS_PLOT'].max()
                    # Aumentamos el "headroom" para que el texto nunca choque con el patrón achurado
                    fig_tot.update_layout(yaxis=dict(range=[0, max_y_tot * 1.35 if max_y_tot > 0 else 1]))
                    
                    for _, row in df_tot_sum.iterrows():
                        if row['CONS_PLOT'] > 0:
                            lbl_val = f"{row['CONS_PLOT']:,.1f}"
                            fig_tot.add_annotation(
                                x=row['CENTRAL'], y=row['CONS_PLOT'], 
                                text=f"<b>Total:<br>{lbl_val}</b>", 
                                showarrow=False, yshift=45, font=dict(size=11)
                            )
                            
                    st.plotly_chart(fig_tot, use_container_width=True)

                    # =====================================================================
                    # ====== NUEVO MÓDULO: ESTIMACIÓN DE DÍAS RESTANTES (PROYECCIÓN) ======
                    # =====================================================================
                    df_stock_proy_base = st.session_state.get('df_stock', pd.DataFrame())
                    if not df_stock_proy_base.empty:
                        df_stk_liq_p = df_stock_proy_base[df_stock_proy_base['TIPO_COMBUSTIBLE'] == comb].copy()
                        if filtro_emp_proy: df_stk_liq_p = df_stk_liq_p[df_stk_liq_p['EMPRESA'].isin(filtro_emp_proy)]
                        if filtro_cen_proy: df_stk_liq_p = df_stk_liq_p[df_stk_liq_p['CENTRAL'].isin(filtro_cen_proy)]
                        
                        if not df_stk_liq_p.empty:
                            st.markdown("---")
                            st.markdown(f"### ⏳ Estimación de Días de Operación Restantes: {comb}")
                            st.info("💡 **Metodología de Cálculo (Estimación Dinámica de Días):**\n"
                                    "Esta métrica proyecta la autonomía operativa de cada central térmica dividiendo el **Inventario de Cierre Real más reciente (IEOD)** "
                                    "entre el **Promedio de Consumo Diario Proyectado** para el rango de fechas seleccionado. "
                                    "A diferencia del cálculo crítico de autonomía (que asume un despacho ininterrumpido a potencia efectiva), "
                                    "esta estimación refleja un escenario operativo realista, basado estrictamente en el comportamiento y tendencias de la programación perfilada de la central.")
                            
                            ultima_fecha_stk_p = df_stk_liq_p['FECHA_OPERATIVA'].max()
                            df_ult_stk_p = df_stk_liq_p[df_stk_liq_p['FECHA_OPERATIVA'] == ultima_fecha_stk_p].copy()
                            df_ult_stk_p['STOCK_PLOT_PROY'] = convertir_volumen(df_ult_stk_p['STOCK_FINAL'], df_ult_stk_p['UNIDADES'], unidad_sel_proy)
                            
                            # Obtenemos el consumo promedio diario a partir de la curva combinada proyectada
                            df_avg_cons = df_combined.groupby('CENTRAL')['CONS_PLOT'].mean().reset_index()
                            df_avg_cons.rename(columns={'CONS_PLOT': 'PROM_DIARIO_PROY'}, inplace=True)
                            
                            df_est_dias = pd.merge(df_ult_stk_p, df_avg_cons, on='CENTRAL', how='inner')
                            df_est_dias['EST_DIAS'] = df_est_dias.apply(
                                lambda row: (row['STOCK_PLOT_PROY'] / row['PROM_DIARIO_PROY']) if row['PROM_DIARIO_PROY'] > 0 else 0,
                                axis=1
                            ).round(2)
                            
                            if not df_est_dias.empty:
                                fig_est_dias = px.bar(
                                    df_est_dias, 
                                    x="CENTRAL", 
                                    y="EST_DIAS", 
                                    color="CENTRAL", 
                                    text_auto='.2f',
                                    custom_data=["STOCK_PLOT_PROY", "PROM_DIARIO_PROY"]
                                )
                                
                                fig_est_dias.update_layout(
                                    height=450, 
                                    xaxis_title="Central Térmica", 
                                    yaxis_title="Días de Operación Estimados",
                                    showlegend=False,
                                    barmode="group"
                                )
                                
                                fig_est_dias.update_traces(
                                    textposition='outside',
                                    hovertemplate=(
                                        "<b>%{x}</b><br>"
                                        "Días Estimados: %{y:.2f} Días<br>"
                                        "Stock Base Cierre: %{customdata[0]:,.2f} " + unidad_sel_proy + "<br>"
                                        "Consumo Promedio Proyectado: %{customdata[1]:,.2f} " + unidad_sel_proy + "/día<extra></extra>"
                                    )
                                )
                                
                                max_y_est = df_est_dias['EST_DIAS'].max()
                                fig_est_dias.update_layout(yaxis=dict(range=[0, max_y_est * 1.25 if max_y_est > 0 else 1]))
                                
                                st.plotly_chart(fig_est_dias, use_container_width=True)

                                # --- Gantt Chart Logic ---
                                st.markdown("#### 📅 Cronograma de Autonomía Proyectada (Gantt)")
                                st.info("Esta gráfica de Gantt ilustra la ventana de tiempo en la que cada central podría seguir operando con el último stock reportado, asumiendo que se mantiene el promedio de consumo proyectado.")
                                
                                df_gantt = df_est_dias.copy()
                                
                                # Filtro específico para el Gantt
                                opciones_gantt = sorted(df_gantt['CENTRAL'].unique())
                                filtro_gantt = st.multiselect(f"🔍 Filtrar Centrales para Gantt ({comb}):", options=opciones_gantt, default=opciones_gantt, key=f"gantt_filtro_{comb}")
                                df_gantt = df_gantt[df_gantt['CENTRAL'].isin(filtro_gantt)]
                                
                                if not df_gantt.empty:
                                    # Comienza en la fecha de inicio de la proyección seleccionada
                                    df_gantt['FECHA_INICIO'] = pd.to_datetime(d_start)
                                    df_gantt['FECHA_FIN'] = df_gantt.apply(lambda row: row['FECHA_INICIO'] + timedelta(days=row['EST_DIAS']), axis=1)
                                    
                                    df_gantt['FECHA_INICIO_STR'] = df_gantt['FECHA_INICIO'].dt.strftime('%d/%m/%Y')
                                    df_gantt['FECHA_FIN_STR'] = df_gantt['FECHA_FIN'].dt.strftime('%d/%m/%Y %H:%M')

                                    # ORDEN DE MENOR A MAYOR AUTONOMÍA
                                    df_gantt = df_gantt.sort_values('FECHA_FIN', ascending=True)
                                    orden_gantt = df_gantt['CENTRAL'].tolist()

                                    fig_gantt = px.timeline(
                                        df_gantt, 
                                        x_start="FECHA_INICIO", 
                                        x_end="FECHA_FIN", 
                                        y="CENTRAL", 
                                        color="CENTRAL",
                                        text="FECHA_FIN_STR",
                                        custom_data=["EST_DIAS", "STOCK_PLOT_PROY", "PROM_DIARIO_PROY", "FECHA_INICIO_STR", "FECHA_FIN_STR"],
                                        category_orders={"CENTRAL": orden_gantt}
                                    )
                                    fig_gantt.update_yaxes(autorange="reversed") # Mostrar de menor a mayor de arriba a abajo
                                    
                                    # MISMO INTERVALO QUE PROYECCIÓN
                                    rango_x_inicio = pd.to_datetime(d_start)
                                    rango_x_fin = pd.to_datetime(d_end) + pd.Timedelta(days=1) # +1 día para incluirlo completamente en la visual
                                    
                                    fig_gantt.update_layout(
                                        height=max(350, len(df_gantt)*45 + 100),
                                        xaxis_title="Ventana de Operación Proyectada",
                                        yaxis_title="Central Térmica",
                                        showlegend=False,
                                        xaxis=dict(range=[rango_x_inicio, rango_x_fin])
                                    )
                                    
                                    fig_gantt.update_traces(
                                        textposition='outside',
                                        hovertemplate=(
                                            "<b>%{y}</b><br>"
                                            "Fecha Inicio (Base): %{customdata[3]}<br>"
                                            "Fecha Fin Estimada: %{customdata[4]}<br>"
                                            "Días de Autonomía: %{customdata[0]:.2f}<br>"
                                            "Stock Restante: %{customdata[1]:,.2f} " + unidad_sel_proy + "<extra></extra>"
                                        )
                                    )
                                    st.plotly_chart(fig_gantt, use_container_width=True)

                                    # Tabla de Gantt
                                    with st.expander("📋 Ver Tabla de Fechas Estimadas de Agotamiento de Stock"):
                                        df_gantt_table = df_gantt[['CENTRAL', 'FECHA_INICIO', 'FECHA_FIN', 'EST_DIAS', 'STOCK_PLOT_PROY', 'PROM_DIARIO_PROY']].copy()
                                        df_gantt_table['FECHA_INICIO'] = df_gantt_table['FECHA_INICIO'].dt.strftime('%d/%m/%Y')
                                        df_gantt_table['FECHA_FIN'] = df_gantt_table['FECHA_FIN'].dt.strftime('%d/%m/%Y %H:%M')
                                        df_gantt_table.columns = ['Central', 'Fecha Inicial de Proyección', 'Fecha Estimada Agotamiento', 'Días Estimados', f'Stock Actual ({unidad_sel_proy})', f'Consumo Promedio ({unidad_sel_proy}/Día)']
                                        st.dataframe(df_gantt_table, use_container_width=True, hide_index=True)

                    # --- TABLAS DE TRAZABILIDAD (UBICADAS AL FINAL) ---
                    st.markdown("---")
                    st.markdown("##### 🗂️ Matriz de Consumo: Evolución Diaria")
                    df_pivot = df_combined.pivot_table(index=['FECHA_OPERATIVA', 'TIPO_DATO'], columns='CENTRAL', values='CONS_PLOT', aggfunc='sum').reset_index()
                    df_pivot.insert(0, 'Día', df_pivot['FECHA_OPERATIVA'].dt.weekday.map(lambda x: DIAS_ESP[x]))
                    df_pivot['FECHA_OPERATIVA'] = df_pivot['FECHA_OPERATIVA'].dt.strftime('%d/%m/%Y')
                    df_pivot = df_pivot.rename(columns={'FECHA_OPERATIVA': 'Fecha', 'TIPO_DATO': 'Estado'}).fillna(0)
                    
                    cols_numericas = [c for c in df_pivot.columns if c not in ['Día', 'Fecha', 'Estado']]
                    st.dataframe(df_pivot.style.format({c: formato_k_m for c in cols_numericas}), use_container_width=True, hide_index=True)

                    with st.expander("🗄️ Ver Matriz Cruda Combinada (IEOD + Estimación)"):
                        df_f_view = df_combined.copy()
                        df_f_view['FECHA_OPERATIVA'] = df_f_view['FECHA_OPERATIVA'].dt.strftime('%d/%m/%Y (%A)')
                        df_f_view.rename(columns={'CONS_PLOT': f'Consumo ({unidad_sel_proy})'}, inplace=True)
                        st.dataframe(df_f_view, use_container_width=True, hide_index=True)
            else:
                st.warning("Selecciona un rango de fechas válido o verifica que haya consumo extraído para operar el modelo predictivo.")
    else:
        st.warning("👈 Por favor, realiza la extracción de datos desde el panel lateral para alimentar el modelo predictivo.")